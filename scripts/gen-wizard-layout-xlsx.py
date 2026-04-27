"""
Genera docs/ux/wizard-layout-analysis.xlsx — analisi layout wizardstep1/2/3.
Estrae dimensioni dai CSS reali (globals.css + componenti) e produce 6 fogli:
  1. step1-grid : sketch visuale wizardstep1 (lista residenze + offers)
  2. step1-data : tabella dimensioni px
  3. step2-grid : sketch visuale wizardstep2 (form + sidebar)
  4. step2-data : tabella dimensioni px
  5. step3-grid : sketch visuale wizardstep3 (single-col)
  6. step3-data : tabella dimensioni px
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

# Colori pastello (coerenti con design system del sito)
C_HEADER = 'd1fae5'   # verde chiaro
C_TITLE  = 'fef3c7'   # giallo chiaro
C_MAIN   = 'dbeafe'   # azzurro chiaro (main wizardstep2)
C_SIDE   = 'fce7f3'   # rosa chiaro (sidebar wizardstep2)
C_GAP    = 'f3f4f6'   # grigio neutro
C_PHOTO  = 'd1d5db'   # grigio (placeholder foto)
C_BANNER_INFO    = 'f0f7ff'
C_BANNER_WARNING = 'fffbeb'
C_BANNER_SUCCESS = 'd1fae5'
C_CTA   = 'FCAF1A'
C_TOTAL = '006CB7'
C_BORDER = '9ca3af'
C_WHITE = 'ffffff'
C_DIVIDER = 'e5e7eb'

FONT_LBL = Font(name='Arial', size=8, color='1f2937')
FONT_LBL_BOLD = Font(name='Arial', size=8, color='1f2937', bold=True)
FONT_LBL_SMALL = Font(name='Arial', size=7, color='6b7280', italic=True)
FONT_TITLE = Font(name='Arial', size=11, color='000000', bold=True)
FONT_TOTAL = Font(name='Arial', size=10, color='ffffff', bold=True)
FONT_CTA = Font(name='Arial', size=10, color='ffffff', bold=True)
FONT_HEAD_TBL = Font(name='Arial', size=10, color='ffffff', bold=True)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN = Side(style='thin', color=C_BORDER)
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_grid_dimensions(ws, n_cols=85, n_rows=110, col_w=2, row_h=15):
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(1, n_rows + 1):
        ws.row_dimensions[r].height = row_h


def draw_block(ws, range_str, fill_color, label='', font=None, merge=True, border=True):
    """Disegna un blocco colorato. Se merge=True, mergea e mette label nel top-left.
       Se merge=False, applica solo fill (nessuna label, tu scriverai in celle interne)."""
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for c in row:
            c.fill = fill
            if border:
                c.border = BOX_BORDER
    if merge:
        ws.merge_cells(range_str)
        top_left = f'{get_column_letter(min_col)}{min_row}'
        cell = ws[top_left]
        if label:
            cell.value = label
        cell.font = font if font else FONT_LBL_BOLD
        cell.alignment = ALIGN_C


def draw_container(ws, range_str, fill_color=None, border=True):
    """Solo bordo + fill di sfondo, NO merge — celle interne rimangono editabili."""
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color) if fill_color else None
    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                               min_col=min_col, max_col=max_col)):
        for col_idx, c in enumerate(row):
            if fill:
                c.fill = fill
            if border:
                # Bordo solo sul perimetro
                left = THIN if col_idx == 0 else None
                right = THIN if col_idx == max_col - min_col else None
                top = THIN if row_idx == 0 else None
                bottom = THIN if row_idx == max_row - min_row else None
                c.border = Border(left=left, right=right, top=top, bottom=bottom)


def write_label(ws, cell_addr, text, font=None, align=None):
    """Scrive un testo in una cella (NON merged)."""
    ws[cell_addr] = text
    ws[cell_addr].font = font if font else FONT_LBL
    if align:
        ws[cell_addr].alignment = align


# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 1: step1-grid (wizardstep1 — lista residenze + offers)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('step1-grid')
set_grid_dimensions(ws, n_cols=85, n_rows=130)

ws['A1'] = 'wizardstep1 — desktop ≥768px (page-container 1200, layout 2-col main 680 + sidebar 380)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:CG1')

# Header sito
draw_block(ws, 'A3:CG4', C_HEADER, 'Header sito (livingapple + nav)  —  ~70px', font=FONT_LBL)

# Stepper (3 step) — niente merge sul container, solo fill
draw_container(ws, 'A6:CG8', C_WHITE, border=False)
write_label(ws, 'AB7', '①  Scegli  (CORRENTE)        ②  Dati ospite        ③  Paga', FONT_LBL_BOLD,
            align=Alignment(horizontal='center', vertical='center'))
ws.merge_cells('AB7:BS7')

# === MAIN (azzurro) cols B:AX (49 col ≈ 680px) ===
draw_block(ws, 'B10:AX10', C_MAIN, '↓ .wizard-container__main  flex 1 · max 680', font=FONT_LBL_BOLD)

# Box riepilogo dates+ospiti+edit (.card-info--compact)
draw_container(ws, 'B12:AX17', C_WHITE, border=True)
write_label(ws, 'C13', '.card-info--compact   →   Riepilogo dates + edit-btn', FONT_LBL_BOLD)
ws.merge_cells('C13:AT13')
write_label(ws, 'C14', '15 mag 2026 – 18 mag 2026   .step1-summary__dates', FONT_LBL_BOLD)
ws.merge_cells('C14:AT14')
write_label(ws, 'C15', '3 notti · 2 adulti   .step1-summary__meta', FONT_LBL)
ws.merge_cells('C15:AT15')
draw_block(ws, 'AU13:AX16', C_BANNER_INFO, '↩\n.step1-\nsummary__\nedit-btn',
           font=Font(name='Arial', size=7, color='006CB7'))

# Titolo
draw_block(ws, 'B19:AX21', C_TITLE,
           'h2.section-title-main  "Scegli la tua casa" / "Scegli la tariffa"', font=FONT_LBL)

# Filter bar
draw_container(ws, 'B23:AX26', C_WHITE, border=True)
write_label(ws, 'C24', '.step1-filter-bar  (overflow-x scroll)', FONT_LBL_BOLD)
ws.merge_cells('C24:AX24')
draw_block(ws, 'C25:M25', C_WHITE, '☰ Filtri  .step1-filter-btn', font=FONT_LBL)
draw_block(ws, 'O25:Y25', C_BANNER_INFO, 'chip filtro  ×', font=FONT_LBL)
draw_block(ws, 'AA25:AK25', C_BANNER_INFO, 'chip filtro  ×', font=FONT_LBL)

# Lista room-card[]
write_label(ws, 'B28', 'LISTA RESIDENZE  →  .step1-room-card[]', FONT_LBL_BOLD)
ws.merge_cells('B28:AX28')

# Card residenza 1
draw_container(ws, 'B30:AX55', C_WHITE, border=True)
write_label(ws, 'C31', '.step1-room-card  (residenza 1: Pink Lady)', FONT_LBL_BOLD)
ws.merge_cells('C31:AX31')

# Card row: photo a sx + details a dx
draw_block(ws, 'C33:N42', C_PHOTO,
           '[FOTO 200×160]\n.step1-room-card__photo-img', font=FONT_LBL)
write_label(ws, 'O33', 'Pink Lady   .step1-room-card__name', FONT_LBL_BOLD)
ws.merge_cells('O33:AW33')
write_label(ws, 'O34', '🏠 App. · 📐 180mq · 👥 12 · 🌊 250m   .step1-room-card__meta-chips', FONT_LBL)
ws.merge_cells('O34:AW34')

# Lista offers dentro la card
write_label(ws, 'O37', 'TARIFFE DISPONIBILI  .step1-room-card__offers', FONT_LBL_SMALL)
ws.merge_cells('O37:AW37')

# Offer 1: Non rimborsabile (selected)
draw_block(ws, 'O39:AW42', C_BANNER_INFO,
           '● Non rimborsabile      79 €/notte   .step1-offer-option.is-selected', font=FONT_LBL)
# Offer 2
draw_block(ws, 'O44:AW47', C_WHITE,
           '○ Parzialmente rimborsabile     85 €/notte   .step1-offer-option', font=FONT_LBL)
# Offer 3
draw_block(ws, 'O49:AW52', C_WHITE,
           '○ Flessibile                    95 €/notte   .step1-offer-option', font=FONT_LBL)

# Card residenza 2 (placeholder)
draw_container(ws, 'B58:AX78', C_WHITE, border=True)
write_label(ws, 'C59', '.step1-room-card  (residenza 2: Casa Verde) — same structure', FONT_LBL_BOLD)
ws.merge_cells('C59:AX59')
draw_block(ws, 'C61:N70', C_PHOTO, '[FOTO 200×160]', font=FONT_LBL)
write_label(ws, 'O61', 'Casa Verde', FONT_LBL_BOLD)
ws.merge_cells('O61:AW61')
write_label(ws, 'O62', '🏠 Villa · 📐 250mq · 👥 8 · 🌊 500m', FONT_LBL)
ws.merge_cells('O62:AW62')
draw_block(ws, 'O66:AW69', C_WHITE, '○ Non rimborsabile      120 €/notte', font=FONT_LBL)
draw_block(ws, 'O71:AW74', C_WHITE, '○ Parzialmente rimborsabile  130', font=FONT_LBL)

# Empty state placeholder note
write_label(ws, 'B82', '... continua con altre residenze (lista scrollabile)', FONT_LBL_SMALL)
ws.merge_cells('B82:AX82')

# Footer dimensione main
write_label(ws, 'B95', '◄── max-width 680 (--container-sm) ──►   flex 1', FONT_LBL_BOLD)
ws.merge_cells('B95:AX95')

# === GAP ===
draw_block(ws, 'AY10:AZ95', C_GAP, '', font=FONT_LBL)

# === SIDEBAR (rosa) — stessa BookingSidebar di step2 ===
draw_block(ws, 'BA10:CA10', C_SIDE, '↓ .wizard-container__sidebar  STICKY 90', font=FONT_LBL_BOLD)

draw_container(ws, 'BA12:CA90', C_WHITE, border=True)
write_label(ws, 'BA13', '.booking-sidebar  width:380  padding:24', FONT_LBL_SMALL)
ws.merge_cells('BA13:CA13')

draw_block(ws, 'BB15:BZ22', C_PHOTO,
           '[ FOTO PINK LADY ]\n.booking-sidebar__hero-img\n332 × 160 (cover)', font=FONT_LBL)
write_label(ws, 'BB24', 'Pink Lady  .section-title-secondary', FONT_LBL_BOLD)
ws.merge_cells('BB24:BZ24')
write_label(ws, 'BB25', 'Appartamento · 180 mq · 12 ospiti', FONT_LBL)
ws.merge_cells('BB25:BZ25')

draw_block(ws, 'BB27:BZ32', C_BANNER_INFO,
           '⚡ Consumi energetici\n.banner.banner--info', font=FONT_LBL)

write_label(ws, 'BB34', 'CARATTERISTICHE  .label-uppercase-muted', FONT_LBL_BOLD)
ws.merge_cells('BB34:BZ34')
write_label(ws, 'BB35', '🛏 4 camere  · 🚿 3 bagni  · 👥 12 ospiti', FONT_LBL)
ws.merge_cells('BB35:BZ35')
write_label(ws, 'BB36', '🏊 Piscina condivisa  ·  🌳 Giardino', FONT_LBL)
ws.merge_cells('BB36:BZ36')

draw_container(ws, 'BB39:BZ42', C_WHITE, border=True)
write_label(ws, 'BB39', 'Date', FONT_LBL_BOLD)
ws.merge_cells('BB39:BZ39')
write_label(ws, 'BB40', '15 mag 2026 – 18 mag 2026', FONT_LBL)
write_label(ws, 'BB41', '3 notti', FONT_LBL_SMALL)

draw_container(ws, 'BB44:BZ46', C_WHITE, border=True)
write_label(ws, 'BB44', 'Ospiti', FONT_LBL_BOLD)
write_label(ws, 'BB45', '2 adulti', FONT_LBL)

write_label(ws, 'BB48', 'DETTAGLI DEL PREZZO  .label-uppercase-muted', FONT_LBL_BOLD)
ws.merge_cells('BB48:BZ48')
write_label(ws, 'BB49', '3 notti × 79 €')
write_label(ws, 'BZ49', '236 €', align=Alignment(horizontal='right'))
write_label(ws, 'BB50', 'Imposta soggiorno')
write_label(ws, 'BZ50', '12 €', align=Alignment(horizontal='right'))

draw_block(ws, 'BB54:BZ56', C_TOTAL, 'Totale                         248 €', font=FONT_TOTAL)

write_label(ws, 'BB58', 'CANCELLAZIONE', FONT_LBL_BOLD)
ws.merge_cells('BB58:BZ58')
write_label(ws, 'BB59', 'Flessibile 60 gg — Cancellazione gratuita', FONT_LBL_SMALL)
ws.merge_cells('BB59:BZ60')

draw_block(ws, 'BB62:BZ69', C_BANNER_WARNING,
           '🛡 Deposito cauzionale — €500\n.banner.banner--warning', font=FONT_LBL)

draw_block(ws, 'BB72:BZ76', C_CTA,
           'CTA  →  "Continua"\n.booking-sidebar__cta\n.btn.btn--primary', font=FONT_CTA)

write_label(ws, 'BB78', 'CIN IT059014B47RVOMN2D  .booking-sidebar__footer', FONT_LBL_SMALL)
ws.merge_cells('BB78:BZ78')

write_label(ws, 'BA92', '◄── 380 px FISSA  flex-shrink:0 ──►', FONT_LBL_BOLD)
ws.merge_cells('BA92:CA92')

# Note sotto
write_label(ws, 'A100', 'NOTE STEP1:', FONT_LBL_BOLD)
write_label(ws, 'A102',
            '• WizardStep1 è il contenuto del .wizard-container__main · main max 680 · sidebar 380 fissa', FONT_LBL)
ws.merge_cells('A102:CG102')
write_label(ws, 'A103',
            '• Lista residenze: .step1-room-card[] — ognuna con foto sx (200×160) + nome + meta-chips + offers', FONT_LBL)
ws.merge_cells('A103:CG103')
write_label(ws, 'A104',
            '• Per ogni residenza: 1-N offers (Non rimborsabile / Parzialmente / Flessibile)', FONT_LBL)
ws.merge_cells('A104:CG104')
write_label(ws, 'A105',
            '• Filter bar: bottone Filtri (apre modale) + chip filtri attivi (×) — overflow-x scroll', FONT_LBL)
ws.merge_cells('A105:CG105')
write_label(ws, 'A106',
            '• Sidebar BookingSidebar: foto Pink Lady + caratteristiche residenza selezionata + Date/Ospiti + Prezzi + CTA Continua', FONT_LBL)
ws.merge_cells('A106:CG106')


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 2: step1-data
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('step1-data')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 55

ws['A1'] = 'wizardstep1 — tabella dimensioni layout (desktop ≥768px, viewport 1200)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

headers = ['Elemento', 'Classe CSS', 'x px', 'y px', 'width px', 'height px', 'parent', 'note']
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

rows_step1 = [
    ('page-container', '.page-container', 0, 0, 1200, 'auto', 'body', 'max-width var(--container-lg) 1200'),
    ('header sito', '(layout root)', 0, 0, 1200, 70, 'page-container', 'sticky'),
    ('stepper', '.ui-stepper', 16, 80, 1168, 60, 'page-container', '3 step (1 corrente, 2 e 3 inactive)'),
    ('wizard-container', '.wizard-container', 0, 140, 1200, 'auto', 'page-container', 'padding lg+base+xl · max 1200'),
    ('layout 2-col', '.wizard-container__layout', 24, 164, 1152, 'auto', 'wizard-container', 'flex row · gap 32'),
    ('main (sx)', '.wizard-container__main', 24, 164, 680, 'auto', 'layout', 'flex 1 · max --container-sm 680'),
    ('  card-summary', '.card-info.card-info--compact', 24, 164, 680, 90, 'main', 'Riepilogo dates+ospiti + edit-btn'),
    ('    dates', '.step1-summary__dates', 40, 180, 600, 22, 'card-summary', '"15 mag – 18 mag" bold 14'),
    ('    meta', '.step1-summary__meta', 40, 205, 600, 18, 'card-summary', '"3 notti · 2 adulti" muted'),
    ('    edit-btn', '.step1-summary__edit-btn', 660, 175, 44, 44, 'card-summary', 'icon undo (touch-target)'),
    ('  titolo', '.section-title-main', 24, 270, 680, 36, 'main', '"Scegli la tua casa" / "Scegli la tariffa"'),
    ('  filter-bar', '.step1-filter-bar', 24, 320, 680, 50, 'main', 'flex · overflow-x scroll · gap 8'),
    ('    filter-btn', '.step1-filter-btn', 24, 330, 100, 40, 'filter-bar', 'apre modale filtri · badge count'),
    ('    filter-chip', '.step1-filter-chip', 132, 332, 90, 36, 'filter-bar', 'chip filtro attivo · close ×'),
    ('  room-card #1', '.step1-room-card', 24, 388, 680, 280, 'main', 'foto sx + name + meta + offers'),
    ('    photo', '.step1-room-card__photo-img', 40, 404, 200, 160, 'room-card', 'rect link · object-fit cover'),
    ('    name', '.step1-room-card__name', 256, 404, 432, 28, 'room-card', '"Pink Lady" 18px / 700'),
    ('    meta-chips', '.step1-room-card__meta-chips', 256, 436, 432, 24, 'room-card', 'icone + numeri (sqm, ospiti, mare, piscina)'),
    ('    offers list', '.step1-room-card__offers', 256, 466, 432, 195, 'room-card', 'lista 1-N offer-option'),
    ('      offer #1 sel', '.step1-offer-option.is-selected', 256, 466, 432, 60, 'offers', 'border 2px primary · bg primary-soft'),
    ('      offer #1 name', '.step1-offer-option__name', 280, 478, 280, 22, 'offer #1', '"Non rimborsabile" bold'),
    ('      offer #1 desc', '.step1-offer-option__desc', 280, 502, 280, 18, 'offer #1', 'descrizione policy'),
    ('      offer #1 price', '.step1-offer-option__price', 580, 478, 100, 22, 'offer #1', '"79 €" 19px / 800 BLU primary'),
    ('      offer #1 /n', '.step1-offer-option__per-night', 580, 502, 100, 16, 'offer #1', '"/notte" 11px muted'),
    ('      offer #2', '.step1-offer-option', 256, 540, 432, 60, 'offers', 'non selezionata · bg bianco'),
    ('      offer #3', '.step1-offer-option', 256, 614, 432, 60, 'offers', 'non selezionata · bg bianco'),
    ('  room-card #2', '.step1-room-card', 24, 686, 680, 280, 'main', 'card residenza 2 — same structure'),
    ('  ...', '...', 24, '...', 680, 'auto', 'main', 'lista scrollabile · 1-N residenze'),
    ('GAP', '(--space-xl 32)', 704, 164, 32, 'auto', 'layout', 'spazio fra main e sidebar'),
    ('sidebar (dx)', '.wizard-container__sidebar', 736, 164, 380, 'auto', 'layout', 'display:none mobile · block 768+'),
    ('  booking-sidebar', '.booking-sidebar', 736, 164, 380, 'auto', 'sidebar', 'sticky 90 · width 380 fissa · padding 24'),
    ('    hero foto', '.booking-sidebar__hero-img', 760, 188, 332, 160, 'booking-sidebar', 'foto residenza selezionata'),
    ('    nome', '.section-title-secondary', 760, 364, 332, 28, 'booking-sidebar', '"Pink Lady"'),
    ('    meta', '.label-metadata', 760, 392, 332, 18, 'booking-sidebar', 'type · sqm · people'),
    ('    banner consumi', '.banner.banner--info', 760, 420, 332, 80, 'booking-sidebar', 'energia info'),
    ('    feature-list', '.booking-sidebar__feature-list', 760, 520, 332, 70, 'booking-sidebar', 'camere, bagni, piscina, giardino'),
    ('    date row', '.booking-sidebar__data-row', 760, 610, 332, 60, 'booking-sidebar', '15 mag – 18 mag · 3 notti'),
    ('    guests row', '.booking-sidebar__data-row', 760, 680, 332, 50, 'booking-sidebar', '2 adulti'),
    ('    price rows', '(.layout-row-between)', 760, 750, 332, 60, 'booking-sidebar', '3 notti × 79 = 236 · imposta 12'),
    ('    totale', '.booking-sidebar__total + __total-value', 760, 820, 332, 30, 'booking-sidebar', 'Totale BLU primary 248 €'),
    ('    cancellazione', '(.label-uppercase-muted + .hint-text)', 760, 860, 332, 50, 'booking-sidebar', 'Flessibile 60gg'),
    ('    banner deposito', '.banner.banner--warning', 760, 920, 332, 90, 'booking-sidebar', 'Deposito €500'),
    ('    CTA Continua', '.booking-sidebar__cta + .btn.btn--primary', 760, 1020, 332, 50, 'booking-sidebar', 'visibile solo se offerta selezionata'),
    ('    CIN', '.booking-sidebar__footer', 760, 1080, 332, 20, 'booking-sidebar', 'CIN IT059...'),
]

for r_idx, row in enumerate(rows_step1):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6, 7) else ALIGN_C
        cell.border = BOX_BORDER
        if isinstance(val, str) and not val.startswith('  ') and c_idx == 0:
            cell.font = FONT_LBL_BOLD

note_row = 4 + len(rows_step1) + 2
notes_step1 = [
    'NOTE STEP1:',
    '• Layout 2-col come step2 (stesso .wizard-container__main 680 + sidebar 380 fissa)',
    '• Main contiene la lista delle residenze · ogni .step1-room-card ha photo sx + name/meta dx + offers sotto',
    '• Sidebar BookingSidebar mostra info residenza SELEZIONATA (cambia al cambio selezione) + CTA "Continua"',
    '• Mobile <768px: sidebar nascosta · stack verticale · CTA Continua dentro la card-room (sticky bottom?)',
    '• filter-bar: overflow-x scroll · bottone Filtri (apre modale bottom-sheet mobile / centered desktop)',
]
for i, n in enumerate(notes_step1):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if i == 0 else FONT_LBL


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 3: step2-grid
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('step2-grid')
set_grid_dimensions(ws, n_cols=85, n_rows=110)

ws['A1'] = 'wizardstep2 — desktop ≥ 1024px (page-container 1200px)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:CG1')

# Header sito
draw_block(ws, 'A3:CG4', C_HEADER, 'Header sito (livingapple + nav)  —  ~70px', font=FONT_LBL)

# Titolo
draw_block(ws, 'A6:CG7', C_TITLE, 'h2.section-title-main  "Conferma e paga"  (22px / 800)', font=FONT_LBL)

# === MAIN (azzurro) cols B:AX (49 col ≈ 680px) ===
draw_block(ws, 'B9:AX9', C_MAIN, '↓ .wizard-step2__main  flex 1 · max 680', font=FONT_LBL_BOLD)

# Contenitore Sezione 1 (no merge, solo bordo + fill bianco)
draw_container(ws, 'C11:AW30', C_WHITE, border=True)
write_label(ws, 'C12', '⓵   .step2-section-card   →   "Come vuoi pagare?"', FONT_LBL_BOLD)
ws.merge_cells('C12:AW12')

# Radio Stripe (selected)
draw_block(ws, 'D14:AV18', C_BANNER_INFO,
           '●  Salva carta, nessun addebito oggi\n.step2-radio-row.is-selected', font=FONT_LBL)
# Radio PayPal
draw_block(ws, 'D20:AV24', C_WHITE,
           '○  Paga in 3 rate [PayPal]\n.step2-radio-row', font=FONT_LBL)
# Banner verde
draw_block(ws, 'D26:AV29', C_BANNER_SUCCESS,
           '✓ Con Stripe la carta viene solo salvata\n.banner.banner--success', font=FONT_LBL)

# Contenitore Sezione 2
draw_container(ws, 'C33:AW65', C_WHITE, border=True)
write_label(ws, 'C34', '⓶   .step2-section-card   →   "I tuoi dati"', FONT_LBL_BOLD)
ws.merge_cells('C34:AW34')

draw_block(ws, 'D36:N39', C_WHITE, '[Nome *]\n.ui-field-input', font=FONT_LBL)
draw_block(ws, 'P36:AV39', C_WHITE, '[Cognome *]\n.ui-field-input', font=FONT_LBL)
draw_block(ws, 'D41:AV44', C_WHITE, '[Email *]', font=FONT_LBL)
draw_block(ws, 'D46:N49', C_WHITE, '[Telefono]', font=FONT_LBL)
draw_block(ws, 'P46:AV49', C_WHITE, '[Paese]', font=FONT_LBL)
draw_block(ws, 'D51:AV54', C_WHITE, '[Ora arrivo --:--]', font=FONT_LBL)
draw_block(ws, 'D56:AV63', C_WHITE,
           '[Richieste speciali — textarea]\n.ui-field-textarea', font=FONT_LBL)

# CTA + terms + back
draw_block(ws, 'B67:AX69', C_CTA,
           'CTA  →  "Continua → 248 €"\n.btn.btn--primary.step2-cta', font=FONT_CTA)
write_label(ws, 'B71', 'Confermando accetti le condizioni generali  (link)  .step2-terms')
ws.merge_cells('B71:AX71')
write_label(ws, 'B73', '← Indietro  .step2-back-link')

# Footer dimensione main
write_label(ws, 'B95', '◄── max-width 680 (--container-sm) ──►   flex 1  min-width 0', FONT_LBL_BOLD)
ws.merge_cells('B95:AX95')

# === GAP ===
draw_block(ws, 'AY9:AZ95', C_GAP, '', font=FONT_LBL)
# Niente label: troppo stretto

# === SIDEBAR (rosa) cols BA:CA (27 col ≈ 380px) ===
draw_block(ws, 'BA9:CA9', C_SIDE, '↓ .wizard-step2__sidebar 380px  STICKY top:90', font=FONT_LBL_BOLD)

# Contenitore booking-sidebar
draw_container(ws, 'BA11:CA94', C_WHITE, border=True)
write_label(ws, 'BA12', '.booking-sidebar  width:380  padding:24', FONT_LBL_SMALL)
ws.merge_cells('BA12:CA12')

# Hero foto
draw_block(ws, 'BB14:BZ21', C_PHOTO,
           '[ FOTO PINK LADY ]\n.booking-sidebar__hero-img\n332 × 160 (cover)', font=FONT_LBL)
# Nome
write_label(ws, 'BB23', 'Pink Lady  .section-title-secondary', FONT_LBL_BOLD)
ws.merge_cells('BB23:BZ23')
# Meta
write_label(ws, 'BB24', 'Appartamento · 180 mq · 12 ospiti  .label-metadata', FONT_LBL)
ws.merge_cells('BB24:BZ24')
# Banner consumi
draw_block(ws, 'BB26:BZ31', C_BANNER_INFO,
           '⚡ Consumi energetici\nbanner.banner--info.banner--with-icon', font=FONT_LBL)
# Voucher
write_label(ws, 'BB33', 'CODICE PROMOZIONALE  .label-uppercase-muted', FONT_LBL)
ws.merge_cells('BB33:BZ33')
draw_block(ws, 'BB34:BS37', C_WHITE, '[es. ESTATE2026]', font=FONT_LBL)
draw_block(ws, 'BU34:BZ37', C_WHITE, 'Applica\n.btn--secondary', font=FONT_LBL)

# Date
draw_container(ws, 'BB39:BZ42', C_WHITE, border=True)
write_label(ws, 'BB39', 'Date  •  Modifica', FONT_LBL_BOLD)
ws.merge_cells('BB39:BZ39')
write_label(ws, 'BB40', '15 mag 2026 – 18 mag 2026', FONT_LBL)
ws.merge_cells('BB40:BZ40')
write_label(ws, 'BB41', '3 notti', FONT_LBL_SMALL)
ws.merge_cells('BB41:BZ41')

# Ospiti
draw_container(ws, 'BB44:BZ46', C_WHITE, border=True)
write_label(ws, 'BB44', 'Ospiti  •  Modifica', FONT_LBL_BOLD)
ws.merge_cells('BB44:BZ44')
write_label(ws, 'BB45', '2 adulti', FONT_LBL)
ws.merge_cells('BB45:BZ45')

# Dettagli prezzo
write_label(ws, 'BB48', 'DETTAGLI DEL PREZZO  .label-uppercase-muted', FONT_LBL_BOLD)
ws.merge_cells('BB48:BZ48')
write_label(ws, 'BB49', '3 notti × 79 €')
write_label(ws, 'BZ49', '236 €', align=Alignment(horizontal='right'))
write_label(ws, 'BB50', 'Imposta di soggiorno')
write_label(ws, 'BZ50', '12 €', align=Alignment(horizontal='right'))
write_label(ws, 'BB51', '€2/pers/notte · max 10 notti', FONT_LBL_SMALL)
ws.merge_cells('BB51:BZ51')

# Servizi extra
write_label(ws, 'BB54', '2. SERVIZI EXTRA (OPZIONALE)', FONT_LBL_BOLD)
ws.merge_cells('BB54:BZ54')
draw_block(ws, 'BB56:BZ59', C_WHITE,
           '🛍 Lettino con biancheria  +40€/u   [-] 0 [+]', font=FONT_LBL)

# Totale
draw_block(ws, 'BB62:BZ64', C_TOTAL,
           'Totale                         248 €', font=FONT_TOTAL)

# Cancellazione
write_label(ws, 'BB66', 'CANCELLAZIONE  .label-uppercase-muted', FONT_LBL_BOLD)
ws.merge_cells('BB66:BZ66')
write_label(ws, 'BB67', 'Flessibile 60 gg — Cancellazione gratuita', FONT_LBL_SMALL)
ws.merge_cells('BB67:BZ68')

# Banner deposito
draw_block(ws, 'BB70:BZ77', C_BANNER_WARNING,
           '🛡 Deposito cauzionale — €500\nbanner.banner--warning.banner--with-icon\nText: alloggio richiede deposito €500...',
           font=FONT_LBL)

# CIN
write_label(ws, 'BB79', 'CIN IT059014B47RVOMN2D  .booking-sidebar__footer', FONT_LBL_SMALL)
ws.merge_cells('BB79:BZ79')

# Footer dimensione sidebar
write_label(ws, 'BA96', '◄── 380 px FISSA  flex-shrink:0  sticky top:90 ──►', FONT_LBL_BOLD)
ws.merge_cells('BA96:CA96')

# Legend
write_label(ws, 'A100', 'LEGENDA COLORI:', FONT_LBL_BOLD)
draw_block(ws, 'A102:F102', C_HEADER, 'Header sito', font=FONT_LBL)
draw_block(ws, 'H102:M102', C_TITLE, 'Titolo h2', font=FONT_LBL)
draw_block(ws, 'O102:T102', C_MAIN, 'Main (action)', font=FONT_LBL)
draw_block(ws, 'V102:AA102', C_SIDE, 'Sidebar', font=FONT_LBL)
draw_block(ws, 'AC102:AH102', C_BANNER_INFO, 'Banner info', font=FONT_LBL)
draw_block(ws, 'AJ102:AO102', C_BANNER_WARNING, 'Banner warning', font=FONT_LBL)
draw_block(ws, 'AQ102:AV102', C_BANNER_SUCCESS, 'Banner success', font=FONT_LBL)
draw_block(ws, 'AX102:BC102', C_CTA, 'CTA giallo', font=FONT_CTA)
draw_block(ws, 'BE102:BJ102', C_TOTAL, 'Totale blu', font=FONT_TOTAL)
draw_block(ws, 'BL102:BQ102', C_PHOTO, 'Foto', font=FONT_LBL)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 2: step2-data
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('step2-data')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 55

ws['A1'] = 'wizardstep2 — tabella dimensioni layout (desktop ≥1024px, viewport 1200)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

headers = ['Elemento', 'Classe CSS', 'x px', 'y px', 'width px', 'height px', 'parent', 'note']
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

rows_step2 = [
    ('page-container', '.page-container', 0, 0, 1200, 'auto', 'body', 'max-width var(--container-lg) 1200, padding-x 16'),
    ('header sito', '(layout root)', 0, 0, 1200, 70, 'page-container', 'sticky top, z-index 10'),
    ('titolo h2', '.section-title-main', 16, 80, 1168, 36, 'page-container', '"Conferma e paga" — text-xl 22 / weight 800'),
    ('layout 2-col', '.wizard-step2__layout', 16, 116, 1168, 'auto', 'page-container', 'flex row · gap 32 (--space-xl) · align-items flex-start'),
    ('main (sx)', '.wizard-step2__main', 16, 116, 680, 'auto', 'layout', 'flex 1 1 0 · min-width 0 · max-width var(--container-sm) 680'),
    ('  sezione 1', '.step2-section-card', 16, 116, 680, 280, 'main', 'Box numerato 1 "Come vuoi pagare?"'),
    ('    radio stripe (selected)', '.step2-radio-row.is-selected', 32, 180, 648, 80, 'sezione 1', 'border 2px primary, bg primary-soft'),
    ('    radio paypal', '.step2-radio-row', 32, 270, 648, 80, 'sezione 1', 'border 1px border, bg bianco'),
    ('    banner stripe info', '.banner.banner--success', 32, 360, 648, 50, 'sezione 1', '"Con Stripe la carta viene solo salvata"'),
    ('  sezione 2', '.step2-section-card', 16, 420, 680, 440, 'main', 'Box numerato 2 "I tuoi dati" (form)'),
    ('    Nome', '.ui-field-input', 32, 480, 320, 48, 'sezione 2', 'col 1 di .step2-form-grid-2'),
    ('    Cognome', '.ui-field-input', 360, 480, 320, 48, 'sezione 2', 'col 2 di .step2-form-grid-2'),
    ('    Email', '.ui-field-input', 32, 540, 648, 48, 'sezione 2', 'full width'),
    ('    Telefono', '.ui-field-input', 32, 600, 320, 48, 'sezione 2', 'col 1'),
    ('    Paese', '.ui-field-input', 360, 600, 320, 48, 'sezione 2', 'col 2'),
    ('    Ora arrivo', '.ui-field-input[type=time]', 32, 660, 648, 48, 'sezione 2', '--:--'),
    ('    Richieste speciali', '.ui-field-textarea', 32, 720, 648, 80, 'sezione 2', 'multi-line'),
    ('  CTA Continua', '.btn.btn--primary.step2-cta', 16, 880, 680, 50, 'main', 'giallo --color-cta · "Continua → 248 €"'),
    ('  terms', '.step2-terms', 16, 940, 680, 20, 'main', 'testo + link "condizioni generali"'),
    ('  back link', '.step2-back-link', 16, 970, 100, 30, 'main', '"← Indietro" grigio secondary'),
    ('GAP', '(--space-xl 32)', 696, 116, 32, 'auto', 'layout', 'spazio fra main e sidebar'),
    ('sidebar (dx)', '.wizard-step2__sidebar', 728, 116, 380, 'auto', 'layout', 'flex-shrink 0 · child width 380 fissa · sticky top 90'),
    ('  booking-sidebar', '.booking-sidebar', 728, 116, 380, 'auto', 'sidebar', 'padding 24 · border 1px · box-shadow sm'),
    ('    hero foto', '.booking-sidebar__hero-img', 752, 140, 332, 160, 'booking-sidebar', 'object-fit cover · radius-md'),
    ('    nome residenza', '.section-title-secondary', 752, 316, 332, 28, 'booking-sidebar', '"Pink Lady" 18px / 700'),
    ('    meta residenza', '.label-metadata', 752, 344, 332, 18, 'booking-sidebar', '"Appartamento · 180 mq · 12 ospiti"'),
    ('    banner consumi', '.banner.banner--info.banner--with-icon', 752, 372, 332, 80, 'booking-sidebar', '"Consumi energetici" full text'),
    ('    voucher row', '(input + .btn--secondary)', 752, 470, 332, 50, 'booking-sidebar', 'input + bottone Applica'),
    ('    date row', '.booking-sidebar__data-row', 752, 540, 332, 70, 'booking-sidebar', 'Date + 3 notti + bottone Modifica'),
    ('    guests row', '.booking-sidebar__data-row', 752, 620, 332, 50, 'booking-sidebar', 'Ospiti + bottone Modifica'),
    ('    price rows', '(.layout-row-between)', 752, 690, 332, 70, 'booking-sidebar', '3 notti × 79 = 236 · imposta 12'),
    ('    extras row', '.wizard-step2-mobile__extra-item', 752, 770, 332, 60, 'booking-sidebar', 'Lettino + counter ±'),
    ('    totale', '.booking-sidebar__total + __total-value', 752, 840, 332, 30, 'booking-sidebar', 'Totale BLU primary 22px / 800 (post Sessione 14)'),
    ('    cancellazione', '.label-uppercase-muted + .hint-text', 752, 880, 332, 50, 'booking-sidebar', '"Flessibile 60 gg — Cancellazione gratuita"'),
    ('    banner deposito', '.banner.banner--warning.banner--with-icon', 752, 940, 332, 90, 'booking-sidebar', '"Deposito cauzionale — €500"'),
    ('    CIN footer', '.booking-sidebar__footer', 752, 1040, 332, 20, 'booking-sidebar', '"CIN IT059014B47RVOMN2D"'),
    ('MARGINE DESTRO', '(spazio libero)', 1108, 116, 92, 'auto', 'page-container', '1200 - 1108 = 92px di "respiro"'),
]

for r_idx, row in enumerate(rows_step2):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6, 7) else ALIGN_C
        cell.border = BOX_BORDER
        if isinstance(val, str) and not val.startswith('  ') and c_idx == 0:
            cell.font = FONT_LBL_BOLD

note_row = 4 + len(rows_step2) + 2
notes = [
    'NOTE LAYOUT:',
    '• 1200 (max-width) - 32 (padding-x 16+16) = 1168 px effettivi disponibili',
    '• 680 (main) + 32 (gap) + 380 (sidebar) = 1092 px usati → 76 px di respiro',
    '• Mobile <768px: sidebar nascosta (display:none) — accordion riepilogo nel main',
    '• Tutti i totali prezzo sono BLU primary post Sessione 14 Fase 2 (553ac03 ratifica)',
]
for i, n in enumerate(notes):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if i == 0 else FONT_LBL


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 3: step3-grid
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('step3-grid')
set_grid_dimensions(ws, n_cols=85, n_rows=80)

ws['A1'] = 'wizardstep3 — desktop (page-container 1200, contenuto centrato 680)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:CG1')

# Header
draw_block(ws, 'A3:CG4', C_HEADER, 'Header sito  ~70px', font=FONT_LBL)

# Margini sx vuoti col A:R, contenuto S:BS, margini dx BT:CG
# Titolo
draw_block(ws, 'S6:BS7', C_TITLE, 'h2.section-title-main  "Conferma prenotazione"', font=FONT_LBL)
write_label(ws, 'S8', '.wizard-step3__subtitle  testo grigio sottotitolo descrittivo', FONT_LBL_SMALL)
ws.merge_cells('S8:BS8')

# Card riepilogo
draw_container(ws, 'S10:BS32', C_WHITE, border=True)
write_label(ws, 'S11', '.card-info  →  CARD RIEPILOGO  (hero+dati+prezzo+totale)', FONT_LBL_BOLD)
ws.merge_cells('S11:BS11')

# Hero compatto
draw_block(ws, 'T13:Y17', C_PHOTO,
           '[FOTO 80×80]\n.wizard-step3__hero-img', font=FONT_LBL)
write_label(ws, 'Z13', 'Pink Lady  .wizard-step3__hero-name', FONT_LBL_BOLD)
ws.merge_cells('Z13:BR13')
write_label(ws, 'Z14', 'Appartamento · 180 mq · 12 ospiti  .wizard-step3__hero-meta', FONT_LBL)
ws.merge_cells('Z14:BR14')

# Divider
draw_block(ws, 'T19:BR19', C_DIVIDER, 'divider .divider-horizontal',
           font=Font(name='Arial', size=6, color='6b7280'))

# Date / Ospiti
write_label(ws, 'T21', 'Date            15 mag 2026 – 18 mag 2026 · 3 notti', FONT_LBL)
ws.merge_cells('T21:BR21')
write_label(ws, 'T22', 'Ospiti          2 adulti', FONT_LBL)
ws.merge_cells('T22:BR22')

draw_block(ws, 'T24:BR24', C_DIVIDER, 'divider',
           font=Font(name='Arial', size=6, color='6b7280'))

# Price rows
write_label(ws, 'T26', 'DETTAGLI PREZZO  .label-uppercase-muted', FONT_LBL_BOLD)
ws.merge_cells('T26:BR26')
write_label(ws, 'T27', '3 notti × 79 €', FONT_LBL)
write_label(ws, 'BO27', '236 €', FONT_LBL, align=Alignment(horizontal='right'))

# Totale
draw_block(ws, 'T29:BR31', C_TOTAL,
           'Totale                                       248 €\n.wizard-step3__total + __total-value (BLU post Sessione 14)',
           font=FONT_TOTAL)

# Banner cancellation
draw_block(ws, 'S34:BS37', C_BANNER_WARNING,
           '📅 Cancellazione gratuita\nbanner.banner--warning  → "Free cancellation up to N days before arrival"',
           font=FONT_LBL)

# Card dati ospite
draw_container(ws, 'S39:BS46', C_WHITE, border=True)
write_label(ws, 'S40', '.card-info  →  DATI OSPITE  (read-only)', FONT_LBL_BOLD)
ws.merge_cells('S40:BS40')
write_label(ws, 'T42', 'DATI OSPITE  .label-uppercase-muted', FONT_LBL_BOLD)
ws.merge_cells('T42:BR42')
write_label(ws, 'T43', 'Luca Rosati  .wizard-step3__guest-name', FONT_LBL_BOLD)
ws.merge_cells('T43:BR43')
write_label(ws, 'T44', 'sistampi@gmail.com  .wizard-step3__guest-meta', FONT_LBL)
ws.merge_cells('T44:BR44')
write_label(ws, 'T45', '+39 ... · IT · arrivo 14:00', FONT_LBL)
ws.merge_cells('T45:BR45')

# Trust signal
draw_block(ws, 'S48:BS49', C_BANNER_SUCCESS,
           '🛡 Pagamento sicuro · Stripe · PayPal\n.wizard-step3__trust', font=FONT_LBL)

# CTA
draw_block(ws, 'S51:BS53', C_CTA,
           '💳 Procedi al pagamento → 248 €\n.btn.btn--primary.wizard-step3__cta', font=FONT_CTA)

# PayPal placeholder
draw_block(ws, 'S55:BS56', 'f0f4f8',
           '⏳ Caricamento PayPal...\n.wizard-step3__paypal-loading (alternativa: PayPal v6 button blue)',
           font=FONT_LBL)

# Back
write_label(ws, 'S58', '← Indietro  .wizard-step3__back-link',
            Font(name='Arial', size=8, color='006CB7', underline='single'))

# Dimension footer
write_label(ws, 'S62', '◄── max-width 680 (--container-sm)  ·  margin auto  centrato ──►', FONT_LBL_BOLD)
ws.merge_cells('S62:BS62')

# Margin labels
write_label(ws, 'A22', 'margine\nsx\n~260 px', Font(name='Arial', size=7, color='6b7280'),
            align=Alignment(horizontal='center', vertical='center', wrap_text=True))
ws.merge_cells('A22:R22')
write_label(ws, 'BT22', 'margine\ndx\n~260 px', Font(name='Arial', size=7, color='6b7280'),
            align=Alignment(horizontal='center', vertical='center', wrap_text=True))
ws.merge_cells('BT22:CG22')


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 4: step3-data
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('step3-data')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 55

ws['A1'] = 'wizardstep3 — tabella dimensioni layout (single-col centrato 680px)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

rows_step3 = [
    ('page-container', '.page-container', 0, 0, 1200, 'auto', 'body', 'max-width 1200, padding-x 16'),
    ('header sito', '(layout root)', 0, 0, 1200, 70, 'page-container', 'sticky'),
    ('wizard-step3', '.wizard-step3', 260, 80, 680, 'auto', 'page-container', 'max-width var(--container-sm) 680 · margin 0 auto · padding 0 var(--space-xs)'),
    ('  titolo', '.section-title-main', 260, 80, 680, 36, 'wizard-step3', '"Conferma prenotazione" — h2 22px / 800'),
    ('  subtitle', '.wizard-step3__subtitle', 260, 120, 680, 24, 'wizard-step3', 'sottotitolo grigio · text-base / muted'),
    ('  CARD riepilogo', '.card-info', 260, 160, 680, 340, 'wizard-step3', 'border 1px · box-shadow sm · padding 16'),
    ('    hero compatto', '.wizard-step3__hero', 276, 180, 648, 80, 'card-riepilogo', 'flex row · align-items center · gap 12'),
    ('      foto 80×80', '.wizard-step3__hero-img', 276, 180, 80, 80, 'hero', 'quadrata · object-fit cover · radius-md'),
    ('      hero-info', '.wizard-step3__hero-info', 368, 180, 556, 80, 'hero', 'flex 1 · min-width 0'),
    ('        hero-name', '.wizard-step3__hero-name', 368, 190, 556, 22, 'hero-info', '"Pink Lady" 18px / 700'),
    ('        hero-meta', '.wizard-step3__hero-meta', 368, 215, 556, 18, 'hero-info', '"type · sqm · people" 13px muted'),
    ('    divider', '.divider-horizontal', 276, 270, 648, 1, 'card-riepilogo', 'hr · color border'),
    ('    data-grid', '.wizard-step3__data-grid', 276, 280, 648, 80, 'card-riepilogo', '<dl> grid 2-col date+ospiti+tariffa'),
    ('    divider', '.divider-horizontal', 276, 365, 648, 1, 'card-riepilogo', ''),
    ('    price-rows', '.wizard-step3__price-rows', 276, 380, 648, 60, 'card-riepilogo', '.layout-row-between · margin-bottom 8'),
    ('    total-row', '.wizard-step3__total-row + __total-value', 276, 450, 648, 30, 'card-riepilogo', 'border-top · totale BLU primary post Sessione 14 (553ac03)'),
    ('  banner cancellation', '.banner.banner--warning.banner--with-icon', 260, 520, 680, 80, 'wizard-step3', 'policy cancellazione (visible solo se cancelPolicy popolato)'),
    ('  CARD dati ospite', '.card-info', 260, 620, 680, 130, 'wizard-step3', 'read-only'),
    ('    label "DATI OSPITE"', '.label-uppercase-muted', 276, 640, 648, 16, 'card-guest', 'uppercase 12px muted'),
    ('    guest-name', '.wizard-step3__guest-name', 276, 660, 648, 24, 'card-guest', 'nome+cognome bold'),
    ('    guest-meta', '.wizard-step3__guest-meta', 276, 690, 648, 18, 'card-guest', 'email'),
    ('    guest-meta extra', '.wizard-step3__guest-meta', 276, 710, 648, 18, 'card-guest', 'phone · country · arrivo'),
    ('  trust signal', '.wizard-step3__trust', 260, 770, 680, 24, 'wizard-step3', 'verde --color-success · "Pagamento sicuro · Stripe · PayPal"'),
    ('  CTA Stripe', '.btn.btn--primary.wizard-step3__cta', 260, 810, 680, 50, 'wizard-step3', 'full-width · giallo brand · "Procedi al pagamento → 248"'),
    ('  PayPal placeholder', '.wizard-step3__paypal-loading', 260, 870, 680, 50, 'wizard-step3', 'alternativa quando paymentMethod === paypal'),
    ('  back link', '.wizard-step3__back-link', 260, 940, 100, 30, 'wizard-step3', '"← Indietro"'),
]

for r_idx, row in enumerate(rows_step3):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6, 7) else ALIGN_C
        cell.border = BOX_BORDER
        if isinstance(val, str) and not val.startswith('  ') and c_idx == 0:
            cell.font = FONT_LBL_BOLD

note_row = 4 + len(rows_step3) + 2
notes_step3 = [
    'NOTE LAYOUT:',
    '• wizardstep3 è SINGLE-COL: max-width 680 (--container-sm) centrato in page-container 1200',
    '• Margine sx ≈ (1200 - 680) / 2 = 260 px · stesso a destra',
    '• Hero compatto 80×80 (NO foto fullwidth come booking-sidebar) — diversa estetica da step2',
    '• Totale prezzo BLU primary (post 553ac03 — coerenza Fase 2 Sessione 14)',
    '• Pagina /paga (PagaClient.tsx) attualmente eredita questo design — feedback utente: vuole adottare invece pattern step2 (2-col)',
]
for i, n in enumerate(notes_step3):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if i == 0 else FONT_LBL


# SAVE
import os
out_path = r'C:\beds24site\docs\ux\wizard-layout-analysis.xlsx'
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f'OK: {out_path} ({os.path.getsize(out_path)} bytes)')
