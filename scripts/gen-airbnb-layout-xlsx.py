"""
Genera docs/ux/airbnb-pattern-analysis.xlsx — analisi layout Airbnb search results.
Dati estratti via Chrome MCP (Claude in Chrome) su www.airbnb.it/s/Milano--Lombardia
viewport 1541×855 (desktop), 18-24 mag 2026, no ospiti.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

C_HEADER = 'd1fae5'
C_TITLE  = 'fef3c7'
C_LIST   = 'dbeafe'
C_MAP    = 'fce7f3'
C_GAP    = 'f3f4f6'
C_PHOTO  = 'd1d5db'
C_BANNER_INFO    = 'f0f7ff'
C_BANNER_WARNING = 'fffbeb'
C_BANNER_SUCCESS = 'd1fae5'
C_CTA   = 'FF385C'   # Airbnb red brand
C_TOTAL = '222222'
C_BORDER = '9ca3af'
C_WHITE = 'ffffff'
C_DIVIDER = 'e5e7eb'

FONT_LBL = Font(name='Arial', size=8, color='1f2937')
FONT_LBL_BOLD = Font(name='Arial', size=8, color='1f2937', bold=True)
FONT_LBL_SMALL = Font(name='Arial', size=7, color='6b7280', italic=True)
FONT_TITLE = Font(name='Arial', size=11, color='000000', bold=True)
FONT_CTA = Font(name='Arial', size=10, color='ffffff', bold=True)
FONT_HEAD_TBL = Font(name='Arial', size=10, color='ffffff', bold=True)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN = Side(style='thin', color=C_BORDER)
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_grid_dimensions(ws, n_cols=110, n_rows=110, col_w=2, row_h=15):
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(1, n_rows + 1):
        ws.row_dimensions[r].height = row_h


def draw_block(ws, range_str, fill_color, label='', font=None, merge=True, border=True):
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for c in row:
            c.fill = fill
            if border:
                c.border = BOX_BORDER
    if merge:
        ws.merge_cells(range_str)
        top_left = f'{get_column_letter(min_col)}{min_row}'
        cell = ws[top_left]
        if label:
            cell.value = label
        cell.font = font if font else FONT_LBL_BOLD
        cell.alignment = ALIGN_C


def draw_container(ws, range_str, fill_color=None, border=True):
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color) if fill_color else None
    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                               min_col=min_col, max_col=max_col)):
        for col_idx, c in enumerate(row):
            if fill:
                c.fill = fill
            if border:
                left = THIN if col_idx == 0 else None
                right = THIN if col_idx == max_col - min_col else None
                top = THIN if row_idx == 0 else None
                bottom = THIN if row_idx == max_row - min_row else None
                c.border = Border(left=left, right=right, top=top, bottom=bottom)


def write_label(ws, cell_addr, text, font=None, align=None):
    ws[cell_addr] = text
    ws[cell_addr].font = font if font else FONT_LBL
    if align:
        ws[cell_addr].alignment = align


# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 1: airbnb-desktop-grid (sketch visuale)
# ═════════════════════════════════════════════════════════════════════════════
# Viewport 1541px → 110 col × 14px ≈ 1540px ✓
ws = wb.create_sheet('airbnb-desktop-grid')
set_grid_dimensions(ws, n_cols=110, n_rows=110)

ws['A1'] = 'Airbnb desktop — search results /s/Milano (viewport misurato 1541×855)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:DF1')

# Header sticky 1526×96, z-index 10 (container fill, niente merge — così disegno dentro)
draw_container(ws, 'A3:DF8', C_HEADER, border=True)
write_label(ws, 'B3', 'Header sticky  1526×96  (z-index 10)', FONT_LBL_BOLD)
ws.merge_cells('B3:K3')

# Logo airbnb (sx) — x≈30, w≈100
draw_block(ws, 'B5:G7', C_WHITE, '🏠 airbnb', font=FONT_LBL_BOLD, border=True)

# Search bar centrata (x=543, w=439, y=25, h=46) — col 40 a col 71
draw_block(ws, 'AN5:BS6', C_WHITE,
           'search bar  439×46  "Milano: alloggi · 18-24 mag · Aggiungi ospiti"',
           font=FONT_LBL, border=True)
draw_block(ws, 'BV5:CB6', C_WHITE, 'Filtri 83×40', font=FONT_LBL, border=True)

# H1 risultati (x=48, y=120, w=239)
write_label(ws, 'D11', '"Oltre 1.000 alloggi: Milano"  H1  239×24  (a y=120 dal viewport)', FONT_LBL_BOLD)
ws.merge_cells('D11:R11')

# === LIST COLUMN (azzurro) cols D:AT (47 col ≈ 658 px) ===
# Più precisamente: x=48 → col 4 (48/14=3.4), w=682 → fino a col 52 (730/14=52)
# Map: x=778 → col 56 (778/14=55.6), w=700 → fino a col 106 (1478/14=105.6)
# Quindi list = D:AZ (col 4..52, 49 col), gap = BA:BC (col 53..55), map = BD:DD (col 56..108)

draw_block(ws, 'D14:AZ14', C_LIST,
           '↓ LIST COLUMN (lista listings) — x=48, w=682, h=5061, y inizia 190',
           font=FONT_LBL_BOLD)

# Card 1 (sx) e Card 2 (dx) — 329×438 ognuna, gap 24 fra
# 329/14 ≈ 23.5 col → uso 24 col per ogni card
# Riga card: y=190 → row ~16 (con scale fudge), height 438px → 22 righe (se row ≈ 20px)
draw_container(ws, 'D16:AA40', C_WHITE, border=True)
write_label(ws, 'D17', 'CARD listing  [itemprop=itemListElement]', FONT_LBL_BOLD)
ws.merge_cells('D17:AA17')
# Photo come container (no merge → posso scrivere ♡ in cella interna)
draw_container(ws, 'E18:Z36', C_PHOTO, border=True)
write_label(ws, 'L26', '[FOTO 329×313]  src=muscache.com/hosting...', FONT_LBL,
            align=Alignment(horizontal='center', vertical='center'))
ws.merge_cells('L26:U28')
# Heart top-right (su photo)
write_label(ws, 'Y18', '♡', FONT_LBL_BOLD, align=Alignment(horizontal='center', vertical='center'))

write_label(ws, 'E37', 'Appartamento ⋅ Milano  (title 22px)', FONT_LBL_BOLD)
ws.merge_cells('E37:Z37')
write_label(ws, 'E38', 'Pellini 3 - 180 mq', FONT_LBL)
ws.merge_cells('E38:Z38')
write_label(ws, 'E39', '3 camere · 6 letti · 1 bagno · Host privato', FONT_LBL_SMALL)
ws.merge_cells('E39:Z39')
write_label(ws, 'E40', '983 € 532 € in totale  (prezzo: vecchio strike + nuovo)', FONT_LBL_BOLD)
ws.merge_cells('E40:Z40')

# Card 2 (a destra)
draw_container(ws, 'AC16:AZ40', C_WHITE, border=True)
write_label(ws, 'AC17', 'CARD listing #2', FONT_LBL_BOLD)
ws.merge_cells('AC17:AZ17')
draw_container(ws, 'AD18:AY36', C_PHOTO, border=True)
write_label(ws, 'AK26', '[FOTO 329×313]', FONT_LBL,
            align=Alignment(horizontal='center', vertical='center'))
ws.merge_cells('AK26:AT28')
write_label(ws, 'AX18', '♡', FONT_LBL_BOLD, align=Alignment(horizontal='center', vertical='center'))
write_label(ws, 'AD37', 'Appartamento · Sesto S. Giovanni', FONT_LBL_BOLD)
ws.merge_cells('AD37:AY37')
write_label(ws, 'AD38', 'Ludo\'s House · Superhost', FONT_LBL)
ws.merge_cells('AD38:AY38')
write_label(ws, 'AD39', '1 camera · 1 letto · Host privato', FONT_LBL_SMALL)
ws.merge_cells('AD39:AY39')
write_label(ws, 'AD40', '524 € 430 € in totale · Cancellazione gratuita', FONT_LBL_BOLD)
ws.merge_cells('AD40:AY40')

# Card 3 e 4 (riga sotto)
draw_container(ws, 'D43:AA67', C_WHITE, border=True)
write_label(ws, 'D44', 'CARD #3', FONT_LBL_BOLD)
ws.merge_cells('D44:AA44')
draw_container(ws, 'E45:Z63', C_PHOTO, border=True)
write_label(ws, 'L53', '[FOTO]', FONT_LBL, align=Alignment(horizontal='center', vertical='center'))

draw_container(ws, 'AC43:AZ67', C_WHITE, border=True)
write_label(ws, 'AC44', 'CARD #4', FONT_LBL_BOLD)
ws.merge_cells('AC44:AZ44')
draw_container(ws, 'AD45:AY63', C_PHOTO, border=True)
write_label(ws, 'AK53', '[FOTO Superhost]', FONT_LBL,
            align=Alignment(horizontal='center', vertical='center'))

# Indicator scroll continua
write_label(ws, 'D70', '...continua infinite scroll · 24 card visibili · scroll-h totale 5940 px', FONT_LBL_SMALL)
ws.merge_cells('D70:AZ70')

# Footer dimensione list column
write_label(ws, 'D72', '◄── LIST COLUMN  682 px wide  (2 card × 329 + gap 24 = 682) ──►', FONT_LBL_BOLD)
ws.merge_cells('D72:AZ72')

# === GAP fra colonne (48 px = 3-4 col) ===
draw_container(ws, 'BA14:BC72', C_GAP, border=False)
write_label(ws, 'BA40', 'gap 48 px', Font(name='Arial', size=7, color='6b7280'),
            align=Alignment(horizontal='center', vertical='center', wrap_text=True))
ws.merge_cells('BA40:BC40')

# === MAP COLUMN (rosa) cols BD:DD (53 col ≈ 700 px), STICKY h=711 ===
draw_block(ws, 'BD14:DD14', C_MAP,
           '↓ MAP COLUMN  x=778, w=700, h=711  ·  position:relative ma sticky comportamento',
           font=FONT_LBL_BOLD)

draw_block(ws, 'BD16:DD55', C_MAP,
           '[ MAPPA Google Maps ]\n[data-testid="map/GoogleMap"]\n\n700 × 711 px\n\nPrezzi in marker (408€, 430€, 538€, 567€, 575€, 598€, ecc.)\n\nIcone: zoom +/- · espansione fullscreen · centratura "Milano" · 2km scale',
           font=FONT_LBL)

write_label(ws, 'BD58', 'Mappa "fissa" mentre la lista scrolla (UX pattern Airbnb signature)', FONT_LBL_BOLD)
ws.merge_cells('BD58:DD58')
write_label(ws, 'BD59', '24 marker prezzo cliccabili · click = scroll alla card corrispondente',
            FONT_LBL_SMALL)
ws.merge_cells('BD59:DD59')

# Footer dimensione map
write_label(ws, 'BD72', '◄── MAP  700 px wide ──►', FONT_LBL_BOLD)
ws.merge_cells('BD72:DD72')

# Margini esterni
write_label(ws, 'A75', 'MARGINI ESTERNI:  sx 48px (col A:C)  ·  dx 48px (col DE:DF)  ·  totale viewport 1541px', FONT_LBL_BOLD)
ws.merge_cells('A75:DF75')

# Note totali
write_label(ws, 'A78', 'CALCOLO LAYOUT:', FONT_LBL_BOLD)
notes_grid = [
    '• viewport: 1541 × 855 (con scrollbar)',
    '• margine sx: 48 + LIST 682 + gap 48 + MAP 700 + margine dx 48 = 1526 (= header.width)',
    '• → margini esterni 48 px simmetrici · gap interno fra colonne 48 px',
    '• LIST: 2 card per riga · 329 + gap 24 + 329 = 682 (esatto)',
    '• MAP: posizione "relative" ma comportamento sticky (resta visibile mentre lista scrolla)',
    '• Card photo: 329 × 313 (aspect 0.95 — quasi quadrata) · NO rounded interno · cuore overlay top-right 32×32',
    '• Card layout interno: photo + (5px) + title + subtitle + meta + prezzo (vecchio strike + nuovo)',
    '• Header sticky z-index 10 · search bar centrata 439×46 · Filtri btn 83×40',
]
for i, n in enumerate(notes_grid):
    cell = ws.cell(row=80 + i, column=1, value=n)
    cell.font = FONT_LBL
    ws.merge_cells(start_row=80 + i, start_column=1, end_row=80 + i, end_column=110)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 2: airbnb-desktop-data
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('airbnb-desktop-data')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 60

ws['A1'] = 'Airbnb desktop — tabella dimensioni layout misurate via Chrome MCP'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

ws['A2'] = 'URL: airbnb.it/s/Milano--Lombardia · viewport 1541×855 · 18-24 mag 2026 · 24 card visibili · scroll totale 5940 px'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:H2')

headers = ['Elemento', 'Selettore / data-testid', 'x px', 'y px', 'width px', 'height px', 'parent', 'note']
for i, h in enumerate(headers):
    c = ws.cell(row=4, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

rows = [
    ('viewport', '(window.innerWidth)', 0, 0, 1541, 855, 'browser', 'scroll-h totale 5940'),
    ('header', 'header (sticky)', 0, 0, 1526, 96, 'viewport', 'position:sticky · z-index:10 · differenza 15px è scrollbar'),
    ('  search bar (centrata)', '[data-testid=little-search] / role=search', 543, 25, 439, 46, 'header', 'campi: località + date + ospiti + lente'),
    ('  filter btn', 'button "Filtri"', 999, 28, 83, 40, 'header', 'apre modale filtri completi'),
    ('main', 'main', 0, 96, 1526, 5291, 'viewport', 'contiene H1 + listColumn + map'),
    ('  H1 risultati', 'h1', 48, 120, 239, 24, 'main', '"Oltre 1.000 alloggi: Milano"'),
    ('  listColumn', '(parent di [itemprop=itemListElement] depth 2)', 48, 190, 682, 5061, 'main', 'colonna SX · 2 card per riga · grid 2-col'),
    ('    card #1', '[itemprop=itemListElement]', 48, 190, 329, 438, 'listColumn', 'prima riga sx'),
    ('      photo', 'img (lazy-load)', 48, 190, 329, 313, 'card #1', 'aspect 0.95 (329/313) · src muscache.com hosting'),
    ('      heart btn', 'button[aria-label*="referit"] overlay', 333, 200, 32, 32, 'card #1', 'pos absolute top-right · salva ai preferiti'),
    ('      title', '[id^="title_"] / h2', 52, 512, 313, 22, 'card #1', '"Appartamento · Milano"'),
    ('      subtitle', '(div meta)', 52, 534, 313, 18, 'card #1', '"Pellini 3 - 180 mq"'),
    ('      meta camere', '(div)', 52, 552, 313, 18, 'card #1', '"3 camere · 6 letti · 1 bagno"'),
    ('      host', '(div)', 52, 570, 313, 18, 'card #1', '"Host privato"'),
    ('      price (strike+new)', '(span)', 52, 609, 143, 18, 'card #1', '"983 € 532 € in totale" · vecchio strike + nuovo'),
    ('    card #2', '[itemprop=itemListElement] [1]', 401, 190, 329, 438, 'listColumn', 'prima riga dx · gap 24 da card #1'),
    ('    card #3 / #4', '[itemprop=itemListElement] [2,3]', '48/401', 658, 329, 438, 'listColumn', 'seconda riga (y=658 = 190+438+30 padding-bottom)'),
    ('    ... 24 card', '[itemprop=itemListElement] [n]', '...', '...', 329, 438, 'listColumn', 'lazy-load infinite scroll'),
    ('GAP cols', '(spazio)', 730, 96, 48, 'auto', 'main', '48 px fra listColumn e map'),
    ('  map', '[data-testid="map/GoogleMap"]', 778, 120, 700, 711, 'main', 'position:relative ma comportamento sticky · 24 marker prezzo'),
    ('    map zoom +', '(google maps controls)', 1342, 200, 32, 32, 'map', 'zoom in'),
    ('    map zoom −', '(google maps controls)', 1342, 232, 32, 32, 'map', 'zoom out'),
    ('    map fullscreen', '(google maps icon)', 1342, 145, 32, 32, 'map', 'expand'),
    ('    price markers', '[role=button][aria-label*="€"]', 'vari', 'vari', 'vari', 'vari', 'map', '24 marker · click → scroll alla card'),
    ('MARGINE DESTRO', '(spazio)', 1478, 96, 48, 'auto', 'viewport', '48 px simmetrici al sx'),
    ('footer', 'footer', 0, 5387, 1526, 'auto', 'viewport', 'mostrato dopo lo scroll della lista (5291 main)'),
]

for r_idx, row in enumerate(rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6, 7) else ALIGN_C
        cell.border = BOX_BORDER
        if isinstance(val, str) and not val.startswith('  ') and c_idx == 0:
            cell.font = FONT_LBL_BOLD

note_row = 5 + len(rows) + 2
notes = [
    'CALCOLI E PRINCIPI UX:',
    '• 1541 = 48 (margine sx) + 682 (lista) + 48 (gap) + 700 (mappa) + 48 (margine dx) + 15 (scrollbar)',
    '• Margini esterni 48 px simmetrici · gap interno colonne 48 px',
    '• LIST 2-col grid: 329 + 24 + 329 = 682 (gap interno cards 24 px · margine inferiore card 30 px)',
    '• Map è "relative" nel DOM, ma viene tenuta in vista via JS scroll listener (effetto sticky)',
    '• Card photo aspect ratio quasi quadrato (0.95) — pattern moderno (era 16:9 in passato)',
    '• Heart btn 32×32 · top-right overlay assoluto · z-index sopra photo',
    '• Prezzo: vecchio prezzo "983 €" strike + nuovo prezzo "532 € in totale" (UX di sconto/promo evidenziato)',
    '• Header sticky z:10 · search bar è "little-search" (compatta), si espande in "big-search" cliccando',
    '• Filtri come bottone separato (no chip pills nel header in questa pagina · sono in modale)',
    '• Nessuna sidebar/categoria visibile (filtri solo dietro bottone)',
    '',
    'CONFRONTO CON wizardstep1 LIVINGAPPLE:',
    '• wizardstep1 attuale: main 680 + sidebar 380 (BookingSidebar fissa con prezzi+CTA) — 2-col asimmetrica',
    '• Airbnb: list 682 + map 700 — 2-col SIMMETRICA (50/50 visivo)',
    '• Differenza chiave: Airbnb mostra MAPPA dx (esplorare geograficamente)  vs  LivingApple mostra SIDEBAR-PREZZO dx (riassumere acquisto)',
    '• Card grid: Airbnb 2-col (4 card per fold visibile)  vs  LivingApple 1-col (1 card per riga, foto sx + offers dx)',
    '• Filtri: Airbnb modale completa  vs  LivingApple bottone + chip orizzontali',
    '• Photo: Airbnb 329×313 quasi quadrata  vs  LivingApple 200×160 rettangolare orizzontale (compatta)',
]
for i, n in enumerate(notes):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if i == 0 or i == 12 else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=8)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 3: searchbar-dropdowns (catturato via Chrome MCP click sui campi)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('searchbar-dropdowns')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 60

ws['A1'] = 'Airbnb searchbar — i 3 dropdown (Dove/Date/Chi) misurati al click'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:G1')

headers_sb = ['Componente', 'Selettore / role', 'x', 'y', 'w', 'h', 'descrizione contenuto']
for i, h in enumerate(headers_sb):
    c = ws.cell(row=3, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

rows_sb = [
    # ── SEARCHBAR COMPATTA (default header) ──
    ('SEARCHBAR compatta', '[data-testid=little-search]', 543, 25, 439, 46,
     '"Milano: alloggi · 18-24 mag · Aggiungi ospiti" + lente rosa · pillola unica'),
    ('  └ Filtri btn (separato)', 'button "Filtri"', 999, 28, 83, 40,
     'separato dalla pillola searchbar · apre la modale full filtri'),

    # ── SEARCHBAR ESPANSA (al click di un campo) ──
    ('SEARCHBAR espansa (3-seg)', '(big-search)', 320, 95, 800, 60,
     'al click campo: pillola si espande in 3 segmenti label-su / value-sotto'),
    ('  └ segmento Dove', '(button con label "Dove")', 320, 95, 240, 60,
     'mostra "Milano, Lombardia" + ×'),
    ('  └ segmento Date', '(button con label "Date")', 580, 95, 240, 60,
     'mostra "18 mag - 24 mag" + ×'),
    ('  └ segmento Chi', '(button con label "Chi")', 840, 95, 240, 60,
     'mostra "4 ospiti, 2 neon..." + bottone Ricerca rosa'),
    ('  └ Ricerca CTA', 'button rosa #FF385C + lente', 1020, 105, 70, 50,
     'icona lente bianca su sfondo rosa · sempre visibile in modalità espansa'),

    # ── DROPDOWN DOVE (suggerimenti località) ──
    ('DROPDOWN Dove', '[role=listbox][aria-label=Suggerimenti di ricerca]',
     346, 204, 398, 360, 'lista 5 suggerimenti località con icona pin · click → riempie input + chiude'),
    ('  └ suggerimento (riga)', '(option role)', 346, 220, 398, 60,
     '<icon-pin> + "Milano, Lombardia" · 5 voci default (storico+near)'),

    # ── DATE PICKER (calendario) ──
    ('CALENDAR Date', '[role=application][aria-label=Calendario]',
     343, 264, 829, 359,
     'calendario 2 mesi affiancati + toggle "Date/Flessibile" + footer check-in/out precision'),
    ('  └ toggle Date/Flessibile', '(tabs)', 575, 200, 440, 56,
     '"Date" (selected pillola bianca) + "Flessibile" — 2 modalità di scelta'),
    ('  └ mese sx (Maggio)', '(table)', 360, 264, 380, 280,
     'griglia 7-col (L M M G V S D) · navigazione frecce sx/dx in alto'),
    ('  └ mese dx (Giugno)', '(table)', 760, 264, 380, 280, 'stessa griglia'),
    ('  └ check-in precision', '(select dropdown)', 470, 600, 250, 50,
     '"Check-in: Giorno esatto / ±1 / ±2 / ±3 / ±7 / ±14 giorni"'),
    ('  └ check-out precision', '(select dropdown)', 720, 600, 250, 50, 'stessa cosa per check-out'),

    # ── STEPPER CHI (ospiti) ──
    ('STEPPER Chi', '(popup absolute)', 750, 180, 400, 380,
     'lista 4 stepper ± con label+sublabel descrittivo'),
    ('  └ Adulti', '(stepper row)', 770, 200, 360, 60,
     '"Adulti" + sublabel "Da 13 in su" · stepper [- 2 +]'),
    ('  └ Bambini', '(stepper row)', 770, 270, 360, 60,
     '"Bambini" + sublabel "Da 2 a 12 anni" · stepper [- 2 +]'),
    ('  └ Neonati', '(stepper row)', 770, 370, 360, 60,
     '"Neonati" + sublabel "Fino a 2 anni" · stepper [- 2 +]'),
    ('  └ Animali domestici', '(stepper row)', 770, 460, 360, 60,
     '"Animali domestici" + link blu "Viaggi con animale di servizio?" · stepper [- 0 +]'),
]

for r_idx, row in enumerate(rows_sb):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6) else ALIGN_C
        cell.border = BOX_BORDER
        if isinstance(val, str) and val.startswith(('SEARCHBAR', 'DROPDOWN', 'CALENDAR', 'STEPPER')) and c_idx == 0:
            cell.font = FONT_LBL_BOLD


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 4: filtri-modale (la modale completa "Filtri")
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('filtri-modale')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 70
ws.column_dimensions['D'].width = 50

ws['A1'] = 'Airbnb modale Filtri — sezioni gerarchiche (568×775 centrata x=487)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:D1')

ws['A2'] = 'Dialogo modale aperto da bottone "Filtri" · 10 sezioni · scroll interno · footer fisso'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:D2')

headers_fm = ['Sezione', 'Pattern UX', 'Contenuto', 'Note']
for i, h in enumerate(headers_fm):
    c = ws.cell(row=4, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

rows_fm = [
    ('Header modale', 'titolo + close X', '"Filtri" centrato + close × top-right', 'sticky in alto'),
    ('1. Consigliati per te', '4 tile-card icona+label',
     '🛏 "2+ camere" · 🧺 "Lavatrice" · 🍴 "Cucina" · 📅 "Cancellazione gratuita"',
     'icone illustrate (non pittogrammi) · click toggle attivo · grid 4-col'),
    ('2. Tipo di alloggio', 'segmented 3-radio',
     '"Qualsiasi tipo" (selected) / "Stanza" / "Casa intera"',
     'pillola unica con border attivo · pattern wizardstep1 nostro filter-chip'),
    ('3. Fascia di prezzo', 'slider min-max + histogram + 2 input',
     'slider 2-handle sopra histogram distribuzione prezzi (rosa Airbnb) · "Minimo €50" / "Massimo €2600+" sotto come input testuali editabili',
     '⭐ pattern unico Airbnb · histogram aiuta scegliere min/max consapevolmente · UX premium'),
    ('4. Stanze e letti', '3 stepper ± con label "Qualsiasi"',
     'Camere da letto [- Qualsiasi +] · Letti [- Qualsiasi +] · Bagni [- Qualsiasi +]',
     'stepper riusa stesso pattern di "Chi/ospiti"'),
    ('5. Servizi', 'checkbox grid (mostra di più)',
     'WiFi · Cucina · A/C · Lavatrice · Asciugatrice · Riscaldamento · TV · ...',
     'inizialmente 8-10 visibili + "Mostra altro" expand · checkbox classici'),
    ('6. Opzioni di prenotazione', '4 chip toggle (con icona)',
     '⚡ "Prenotazione immediata" · 🔑 "Self check-in" · 📅 "Cancellazione gratuita" · 🐾 "Animali ammessi"',
     'chip pillola con icona linea + label · click toggle · grid 2x2 desktop'),
    ('7. Alloggi straordinari', '2 tile-card grandi',
     '"Amato dagli ospiti" (Gli alloggi più amati su Airbnb) · "Luxe" (Alloggi di lusso dal design favoloso)',
     'tile più grandi delle "Consigliati" · 2-col · sublabel descrittivo · brand-led'),
    ('8. Tipo di alloggio (accordion)', 'accordion ▼',
     'Appartamento / Loft / Villa / Cottage / Stanza / B&B / ...',
     'collapsed default · expand mostra checkbox grid'),
    ('9. Caratteristiche accessibilità', 'accordion ▼',
     'Ingresso senza gradini · Bagno accessibile in sedia a rotelle · ...',
     'collapsed default · WCAG-aware'),
    ('10. Lingua dell\'host', 'accordion ▼',
     'Italiano / English / Deutsch / ...',
     'collapsed default'),
    ('Footer fisso', '2 bottoni',
     '"Cancella tutto" (sx, link blu)  +  "Mostra oltre 1.000 alloggi" (dx, primary rosa #FF385C)',
     '⭐ il count si aggiorna LIVE mentre selezioni filtri (feedback immediato)'),
]

for r_idx, row in enumerate(rows_fm):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L
        cell.border = BOX_BORDER
        if c_idx == 0:
            cell.font = FONT_LBL_BOLD

# Note finali con UX pattern e confronto
note_row = 5 + len(rows_fm) + 2
notes_fm = [
    'PATTERN UX CHIAVE (Airbnb Filtri):',
    '• Mai uniform-pattern: ogni sezione ha il pattern UX più adatto al tipo di filtro (slider per range, stepper per count, segmented per esclusivi, chip toggle per non-mutex, accordion per long lists)',
    '• "Consigliati per te" in cima · personalizzato (più cliccati / più rilevanti per la query)',
    '• Histogram prezzo è SIGNATURE: aiuta capire dove sta la massa dei prezzi → scelta informata',
    '• Footer LIVE count: "Mostra oltre 1.000 alloggi" si aggiorna ad ogni cambio · feedback immediato di rilevanza',
    '• Accordion in fondo per filtri "long-tail" (tipo alloggio specifico, accessibilità, lingua) · evita modale infinita',
    '• Nessun click "Applica" intermedio · TUTTO live · solo "Cancella tutto" e "Mostra"',
    '',
    'CONFRONTO CON wizardstep1 LIVINGAPPLE:',
    '• LivingApple: bottone Filtri + chip orizzontali sopra la lista · modale bottom-sheet mobile / centered desktop',
    '• Airbnb: bottone Filtri solo · modale ricca · chip non in pagina (filtri appaiono solo in modale)',
    '• LivingApple ha 5 macro-filtri (Sort, Mare, Piscina, Tipo, Camere) · Airbnb ne ha 10 · scopo diverso (booking vs platform)',
    '• Pattern Airbnb da adottare: histogram prezzo + LIVE count footer + tile-card "Consigliati"',
    '• Pattern LivingApple da MANTENERE: chip orizzontali above-fold per filtri primari (più veloce di aprire modale)',
]
for i, n in enumerate(notes_fm):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if (i == 0 or i == 8) else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=4)


# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
import os
out_path = r'C:\beds24site\docs\ux\airbnb-pattern-analysis.xlsx'
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f'OK: {out_path} ({os.path.getsize(out_path)} bytes)')
