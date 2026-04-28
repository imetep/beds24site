"""
Aggiorna docs/ux/airbnb-pattern-analysis.xlsx con 4 fogli nuovi:
  - breakpoints       : 11 viewport width misurate (3440 → 390) con cambi layout
  - airbnb-mobile-grid : sketch visuale iPhone 378×819 (UA mobile)
  - airbnb-mobile-data : tabella dimensioni mobile
  - airbnb-mobile-flow : flow searchbar progressive (Dove→Quando→Chi) + Filtri fullscreen + drag handle
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

C_HEADER = 'd1fae5'
C_TITLE  = 'fef3c7'
C_LIST   = 'dbeafe'
C_MAP    = 'fce7f3'
C_GAP    = 'f3f4f6'
C_PHOTO  = 'd1d5db'
C_BANNER_INFO    = 'f0f7ff'
C_BANNER_WARNING = 'fffbeb'
C_BANNER_SUCCESS = 'd1fae5'
C_CTA   = 'FF385C'
C_CTA_BLACK = '222222'
C_BORDER = '9ca3af'
C_WHITE = 'ffffff'
C_DIVIDER = 'e5e7eb'

FONT_LBL = Font(name='Arial', size=8, color='1f2937')
FONT_LBL_BOLD = Font(name='Arial', size=8, color='1f2937', bold=True)
FONT_LBL_SMALL = Font(name='Arial', size=7, color='6b7280', italic=True)
FONT_TITLE = Font(name='Arial', size=11, color='000000', bold=True)
FONT_CTA = Font(name='Arial', size=10, color='ffffff', bold=True)
FONT_HEAD_TBL = Font(name='Arial', size=10, color='ffffff', bold=True)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN = Side(style='thin', color=C_BORDER)
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_grid_dimensions(ws, n_cols=110, n_rows=110, col_w=2, row_h=15):
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(1, n_rows + 1):
        ws.row_dimensions[r].height = row_h


def draw_block(ws, range_str, fill_color, label='', font=None, merge=True, border=True):
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for c in row:
            c.fill = fill
            if border:
                c.border = BOX_BORDER
    if merge:
        ws.merge_cells(range_str)
        top_left = f'{get_column_letter(min_col)}{min_row}'
        cell = ws[top_left]
        if label:
            cell.value = label
        cell.font = font if font else FONT_LBL_BOLD
        cell.alignment = ALIGN_C


def draw_container(ws, range_str, fill_color=None, border=True):
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color) if fill_color else None
    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row,
                                               min_col=min_col, max_col=max_col)):
        for col_idx, c in enumerate(row):
            if fill:
                c.fill = fill
            if border:
                left = THIN if col_idx == 0 else None
                right = THIN if col_idx == max_col - min_col else None
                top = THIN if row_idx == 0 else None
                bottom = THIN if row_idx == max_row - min_row else None
                c.border = Border(left=left, right=right, top=top, bottom=bottom)


def write_label(ws, cell_addr, text, font=None, align=None):
    ws[cell_addr] = text
    ws[cell_addr].font = font if font else FONT_LBL
    if align:
        ws[cell_addr].alignment = align


# Carica xlsx esistente o crea nuovo
import os
out_path = r'C:\beds24site\docs\ux\airbnb-pattern-analysis.xlsx'
if os.path.exists(out_path):
    wb = load_workbook(out_path)
else:
    wb = Workbook()
    wb.remove(wb.active)

# Rimuovi fogli con stesso nome se esistono (per overwrite clean)
for sheet_name in ['breakpoints', 'airbnb-mobile-grid', 'airbnb-mobile-data', 'airbnb-mobile-flow']:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO: breakpoints (11 width misurate)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('breakpoints')
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 60

ws['A1'] = 'Airbnb — 11 breakpoint misurati via Chrome MCP (DevTools Responsive + iPhone 14 Pro)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:G1')

headers = ['Width (px)', 'Map state', 'List grid', 'Header pattern', 'Searchbar w', 'Filter btn', 'Note']
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

bp_rows = [
    ('2408', 'visible dx 1125w', '3 cards/riga (359w)', 'searchbar centrata 388×46', '388', '83×40', 'Layout XL ultrawide · listCol 1124'),
    ('1920', 'visible dx 885w', '2 cards/riga (426w)', 'searchbar centrata 388×46', '388', '83×40', 'Layout L · listCol 876'),
    ('1541', 'visible dx 700w', '2 cards/riga (329w)', 'searchbar centrata 388×46', '388', '83×40', 'Desktop standard (riferimento)'),
    ('1440', 'visible dx 650w', '2 cards/riga (303w)', 'searchbar centrata 388×46', '388', '83×40', 'Laptop standard'),
    ('1280', 'visible dx 433w', '2 cards/riga (332w)', 'searchbar INLINE LEFT (vicino logo)', '388', '83×40', '🎯 Header layout change (centrata→inline)'),
    ('1024', 'visible dx 473w', '1 card/riga (440w)', 'searchbar inline-left', '388', '83×40', '🎯 Card grid 2→1 col (in modo desktop)'),
    ('950',  'visible dx 399w', '1 card/riga (440w)', 'searchbar inline-left', '388', '83×40', 'Mostra mappa btn pre-cooked nel DOM'),
    ('800',  'HIDDEN + floating btn "Mostra la mappa" (173×48 bottom-center)', '2 cards/riga (349w)', 'searchbar inline-left', '388', '83×40', '🎯 Map breakpoint (mappa sparisce, lista riempie)'),
    ('743',  'FULL-WIDTH TOP + drag handle linea grigia centrata', '2 cards/riga (328w)', 'searchbar PILLOLA mobile + back arrow + filtri icon', 'no little-search', 'icon 40×40', '🎯 Mobile mode (bottom nav: Esplora·Preferiti·Viaggi·Messaggi·Profilo)'),
    ('600',  'full-width top', '2 cards/riga (257w)', 'searchbar pillola mobile', 'no little-search', 'icon 40×40', 'Mobile mid-range'),
    ('390',  'full-width top (363×819 con UA iPhone)', '1 card/riga (315w)', 'searchbar pillola "Milano: alloggi" + ← back + ≡ filtri', 'no little-search', 'icon 40×40', '🎯 Card grid 2→1 mobile (iPhone 14 Pro UA mobile)'),
]
for r_idx, row in enumerate(bp_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (1, 2, 3, 6) else ALIGN_C
        cell.border = BOX_BORDER
        # Highlight i breakpoint critici
        if isinstance(val, str) and ('🎯' in val or 'HIDDEN' in val or 'PILLOLA' in val or 'INLINE LEFT' in val):
            cell.fill = PatternFill('solid', start_color='fef3c7', end_color='fef3c7')

# Note sotto tabella
note_row = 4 + len(bp_rows) + 2
notes = [
    'BREAKPOINT CRITICI IDENTIFICATI:',
    '• ~2200 px : 3→2 cards/riga (in modalità desktop XL)',
    '• ~1440 px : searchbar centrata → inline-left (header layout change)',
    '• ~1100 px : 2→1 card/riga (modalità desktop con sidebar mappa)',
    '• ~870 px : mappa scompare → "Mostra mappa" floating btn appare',
    '• ~770 px : layout desktop → mobile completo (mappa torna full-top, drag handle, pillola searchbar, bottom nav)',
    '• ~480 px : in mobile, 2→1 card/riga',
    '',
    'CONFRONTO con LivingApple wizardstep1 (CSS audit):',
    '• Airbnb usa 6 breakpoint visibili, LivingApple ne ha 3 (`--bp-sm 640 / --bp-md 768 / --bp-lg 1024`)',
    '• Airbnb sezionalizza in 4 zone: ultrawide (>2200), desktop (1100-2200), tablet (770-1100), mobile (<770)',
    '• LivingApple ha solo desktop (≥768) vs mobile (<768) — molto più semplificato',
    '• Per redesign wizardstep1: aggiungere breakpoint intermedio ~1100 per cambiare grid sidebar BookingSidebar?',
]
for i, n in enumerate(notes):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if i == 0 or i == 8 else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=7)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO: airbnb-mobile-grid (sketch iPhone 378×819)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('airbnb-mobile-grid')
# 378 px / 14 ≈ 27 col → uso 28 col larghe
set_grid_dimensions(ws, n_cols=28, n_rows=80, col_w=2, row_h=15)

ws['A1'] = 'Airbnb iPhone 14 Pro (378×819, UA mobile, DPR 1.25)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:AB1')

# Header pillola (back + Milano + filtri)
draw_container(ws, 'A3:AB5', C_WHITE, border=True)
draw_block(ws, 'A3:B5', C_WHITE, '←', font=FONT_LBL_BOLD, border=True)
draw_block(ws, 'D3:Y5', C_WHITE, 'Milano: alloggi\n18-24 mag · 4 ospiti', font=FONT_LBL, border=True)
draw_block(ws, 'AA3:AB5', C_WHITE, '≡', font=FONT_LBL_BOLD, border=True)
write_label(ws, 'A6', 'header pillola 40×40 sx (back) + 211×58 centro (searchbar) + 40×40 dx (filtri)', FONT_LBL_SMALL)
ws.merge_cells('A6:AB6')

# Mappa fullwidth top
draw_block(ws, 'A8:AB30', C_MAP, '[ MAPPA Google Maps full-width ]\n363 × 788 px\n\nPrice markers (760€ · 891€ · 1.402€ · 1.036€ · 1.035€ · 859€ · 813€ · ecc.)\n\n📍 Milano centro pin\n2 km scale', font=FONT_LBL)

# Drag handle (linea grigia centrata)
draw_block(ws, 'L31:Q31', C_BORDER, '═══', font=Font(name='Arial', size=8, color='6b7280', bold=True), border=False)
write_label(ws, 'A32', '↕ DRAG HANDLE (linea grigia centrata · 157×0px coord, visibile come pill 30×4 pattern bottom-sheet) ↕', FONT_LBL_SMALL)
ws.merge_cells('A32:AB32')

# Drawer content
draw_container(ws, 'A33:AB67', C_WHITE, border=True)
write_label(ws, 'A34', 'DRAWER (pull-up bottom-sheet · default peek-bottom)', FONT_LBL_BOLD)
ws.merge_cells('A34:AB34')
write_label(ws, 'A35', 'Oltre 1.000 alloggi  (h1)', FONT_LBL_BOLD,
            align=Alignment(horizontal='center'))
ws.merge_cells('A35:AB35')
write_label(ws, 'A36', 'Come ordiniamo i risultati ⓘ', FONT_LBL_SMALL,
            align=Alignment(horizontal='center'))
ws.merge_cells('A36:AB36')

# Card listing 1-col
draw_container(ws, 'B38:AA62', C_WHITE, border=True)
write_label(ws, 'B39', 'CARD listing (315×451 px · 1 per riga)', FONT_LBL_BOLD)
ws.merge_cells('B39:AA39')
draw_container(ws, 'C40:Z58', C_PHOTO, border=True)
write_label(ws, 'L48', '[ FOTO 315×395 ]', FONT_LBL,
            align=Alignment(horizontal='center', vertical='center'))
ws.merge_cells('L48:Q48')
write_label(ws, 'X40', '♡', FONT_LBL_BOLD,
            align=Alignment(horizontal='center', vertical='center'))
write_label(ws, 'C59', 'Superhost / Amato dagli ospiti  (badge top-left photo)', FONT_LBL_SMALL)
ws.merge_cells('C59:Z59')
write_label(ws, 'C60', 'Appartamento · Milano  (title 22px bold)', FONT_LBL_BOLD)
ws.merge_cells('C60:Z60')
write_label(ws, 'C61', 'Pellini 3 · 180mq · 4.93 (248)', FONT_LBL)
ws.merge_cells('C61:Z61')
write_label(ws, 'C62', '983€ 532€ in totale  (prezzo strike+nuovo)', FONT_LBL_BOLD)
ws.merge_cells('C62:Z62')

# Bottom nav fixed
draw_container(ws, 'A69:AB73', C_WHITE, border=True)
write_label(ws, 'A69', '◄── BOTTOM NAV (sticky position:fixed bottom · 5 voci icon+label) ──►', FONT_LBL_BOLD,
            align=Alignment(horizontal='center'))
ws.merge_cells('A69:AB69')
draw_block(ws, 'A71:E72', C_WHITE, '🔍\nEsplora\n(SEL)', font=Font(name='Arial', size=7, color='FF385C', bold=True), border=True)
draw_block(ws, 'F71:K72', C_WHITE, '♡\nPreferiti', font=FONT_LBL, border=True)
draw_block(ws, 'L71:Q72', C_WHITE, '🏠\nViaggi', font=FONT_LBL, border=True)
draw_block(ws, 'R71:V72', C_WHITE, '💬\nMessaggi\n•', font=FONT_LBL, border=True)
draw_block(ws, 'W71:AB72', C_WHITE, '👤\nProfilo', font=FONT_LBL, border=True)

# Note finali
write_label(ws, 'A75', 'PATTERN UX MOBILE chiave:', FONT_LBL_BOLD)
ws.merge_cells('A75:AB75')
write_label(ws, 'A76', '• Map fullscreen-top + drawer pull-up (drag handle decorativo, drag verticale = expand)', FONT_LBL)
ws.merge_cells('A76:AB76')
write_label(ws, 'A77', '• Header MAI mostra logo airbnb (solo la pillola contesto-corrente)', FONT_LBL)
ws.merge_cells('A77:AB77')
write_label(ws, 'A78', '• Card 1 col, foto quasi quadrata, prezzo prominente (CTA conversion)', FONT_LBL)
ws.merge_cells('A78:AB78')
write_label(ws, 'A79', '• Bottom nav sticky 5 voci (Esplora=current rosa #FF385C, altri grigi)', FONT_LBL)
ws.merge_cells('A79:AB79')


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO: airbnb-mobile-data
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('airbnb-mobile-data')
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 60

ws['A1'] = 'Airbnb iPhone 14 Pro — tabella dimensioni misurate'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

ws['A2'] = 'Viewport 378×819 · DPR 1.25 · UA: Mozilla/5.0 (iPhone; CPU iPhone OS 18_5) · 24 card lazy-loaded'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:H2')

headers = ['Elemento', 'Selettore / aria-label', 'x px', 'y px', 'width px', 'height px', 'parent', 'note']
for i, h in enumerate(headers):
    c = ws.cell(row=4, column=i + 1, value=h)
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

mobile_rows = [
    ('viewport', 'window.innerWidth', 0, 0, 378, 819, 'browser', 'iPhone 14 Pro · DPR 1.25'),
    ('header pillola', 'button[aria-label="Milano: alloggi"]', 76, 12, 211, 58, 'viewport', 'pillola compatta · click → modale searchbar fullscreen accordion'),
    ('  back icon', 'button[aria-label="Indietro"]', 24, 21, 40, 40, 'header', 'top-left · ← ritorna a home'),
    ('  filtri icon', 'button[aria-label="Mostra i filtri"]', 299, 21, 40, 40, 'header', 'top-right · ≡ → modale Filtri fullscreen'),
    ('map', '[data-testid="map/GoogleMap"]', 0, 0, 363, 788, 'main', 'fullscreen TOP · 24 marker prezzo · click marker = scroll card'),
    ('drag handle', '(div h≈4 centrata)', 103, 500, 157, 0, 'main', 'pill grigia · drag verticale = espande/riduce drawer (pattern bottom-sheet)'),
    ('drawer (default)', '(parent del card list)', 0, 515, 378, 'auto', 'main', 'pull-up · peek-bottom default · contiene heading + sort + lista cards'),
    ('  heading', 'h1 "Oltre 1.000 alloggi"', 0, 580, 378, 22, 'drawer', 'centrato'),
    ('  sort label', '"Come ordiniamo i risultati"', 0, 605, 378, 18, 'drawer', 'link con info icon'),
    ('  card #1', '[itemprop="itemListElement"]', 24, 515, 315, 451, 'drawer', '1 per riga · foto + meta + prezzo'),
    ('    photo', 'img', 24, 515, 315, 395, 'card', 'aspect 0.80 (315/395) · object-fit cover · radius'),
    ('    superhost badge', '(span overlay top-left)', 36, 530, 90, 28, 'photo', 'optional · "Superhost" / "Amato dagli ospiti" / vuoto'),
    ('    heart btn', 'button (top-right photo)', 287, 525, 32, 32, 'photo', '♡ salva preferiti'),
    ('    title', '"Appartamento · Milano"', 24, 920, 315, 22, 'card', 'h2 22px / 700'),
    ('    meta', '"Pellini 3 · 180mq · ★4.93 (248)"', 24, 942, 315, 18, 'card', 'rating + meta su singola riga (compatto)'),
    ('    price', '"983€ 532€ in totale"', 24, 962, 315, 22, 'card', 'strike + nuovo · bold'),
    ('bottom nav', '(div fixed bottom 5 voci)', 0, 771, 378, 48, 'viewport', 'STICKY · 5 voci icon+label'),
    ('  Esplora (current)', '(rosa #FF385C)', 0, 771, 76, 48, 'bottom-nav', '🔍 + label · current state'),
    ('  Preferiti', '(grigio default)', 76, 771, 76, 48, 'bottom-nav', '♡ + label'),
    ('  Viaggi', '(grigio default)', 152, 771, 76, 48, 'bottom-nav', '🏠 logo airbnb mini + label'),
    ('  Messaggi', '(grigio + dot rosso)', 228, 771, 76, 48, 'bottom-nav', '💬 + label · badge unread (•)'),
    ('  Profilo', '(grigio default)', 304, 771, 76, 48, 'bottom-nav', '👤 + label'),
]

for r_idx, row in enumerate(mobile_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6, 7) else ALIGN_C
        cell.border = BOX_BORDER
        if isinstance(val, str) and not val.startswith('  ') and c_idx == 0:
            cell.font = FONT_LBL_BOLD


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO: airbnb-mobile-flow (searchbar accordion + Filtri)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('airbnb-mobile-flow')
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 60

ws['A1'] = 'Airbnb iPhone — Flow searchbar accordion + Filtri fullscreen + drag handle'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:C1')

ws['A3'] = 'SEARCHBAR MOBILE — Modale fullscreen ACCORDION (3 step progressive)'
ws['A3'].font = Font(name='Arial', size=11, color='000000', bold=True)
ws['A3'].fill = PatternFill('solid', start_color='dbeafe', end_color='dbeafe')
ws.merge_cells('A3:C3')

ws['A4'] = 'Trigger: tap sulla pillola searchbar nel header. Apre modale fullscreen 378×819 con tabs 🏠Alloggi / 🎈Esperienze / 🛎️Servizi in alto.'
ws['A4'].font = FONT_LBL
ws.merge_cells('A4:C4')

flow_rows = [
    # (step, contenuto, footer CTA)
    ('1️⃣ Dove? (espansa di default)',
     'Search input "Milano" + Lista "Ricerche recenti" (Milano · 18-24 mag, Scauri · 11-18 mag) + "Destinazioni suggerite" (📍Nelle vicinanze, 🏙️Milano Lombardia "Ideale per famiglie", 🇫🇷Parigi…)',
     'Footer: "Cancella tutto" sx + 🔍 "Ricerca" rosa CTA dx (oppure tap "Quando" collapsed sotto per andare avanti)'),
    ('2️⃣ Quando?',
     'Toggle segmented "Date | Date flessibili" + Calendar 1 mese alla volta (vs 2 mesi affiancati desktop) + "Carica date precedenti" btn + Footer dropdowns "Check-in/Check-out: Giorno esatto" (precision flexibility)',
     'Footer: "Ripristina" sx + "Avanti" CTA NERO dx (step-progression, NO Ricerca qui)'),
    ('3️⃣ Chi?',
     'Stepper ± per 4 categorie: Adulti (Da 13 in su) · Bambini (Da 2 a 12 anni) · Neonati (Fino a 2 anni) · Animali domestici (link "Viaggi con un animale di servizio?")',
     'Footer: "Cancella tutto" sx + 🔍 "Ricerca" rosa CTA dx (CTA finale, lancia la ricerca)'),
    ('🔄 Backtracking',
     'Sezioni precedenti restano visibili come collapsed cards in cima (mostrano "Dove · Milano, Lombardia" / "Quando · 18 mag - 24 mag"). Tap su una di esse riapre quella sezione.',
     '—'),
    ('❌ Chiudi',
     'Top-right: bottone X tondo per chiudere modale senza applicare modifiche. Comportamento: torna a pagina lista cards.',
     '—'),
]

ws['A6'] = 'Step'
ws['B6'] = 'Contenuto'
ws['C6'] = 'Footer / CTA'
for c in [ws['A6'], ws['B6'], ws['C6']]:
    c.font = FONT_HEAD_TBL
    c.fill = PatternFill('solid', start_color='374151', end_color='374151')
    c.alignment = ALIGN_C
    c.border = BOX_BORDER

for r_idx, row in enumerate(flow_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=7 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L
        cell.border = BOX_BORDER
        cell.fill = PatternFill('solid', start_color='f9fafb', end_color='f9fafb')

# Sezione Filtri
fr = 7 + len(flow_rows) + 2
ws.cell(row=fr, column=1, value='FILTRI MOBILE — Modale fullscreen (NON bottom-sheet)').font = Font(name='Arial', size=11, color='000000', bold=True)
ws.cell(row=fr, column=1).fill = PatternFill('solid', start_color='fef3c7', end_color='fef3c7')
ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=3)

ws.cell(row=fr+1, column=1, value='Trigger: tap icona ≡ filtri (top-right header). Modale fullscreen 378×~auto (scroll interno).').font = FONT_LBL
ws.merge_cells(start_row=fr+1, start_column=1, end_row=fr+1, end_column=3)

filter_rows = [
    ('Header', 'Titolo "Filtri" centrato + ✕ chiudi top-right', '—'),
    ('1. Consigliati per te', '3 tile-card (vs 4 desktop): icone illustrate (2+ camere · Culla · Lavatrice)', 'Adapt per mobile width'),
    ('2. Tipo di alloggio', 'Segmented 3 (Qualsiasi tipo / Stanza / Casa intera)', 'Identico desktop'),
    ('3. Fascia di prezzo', 'Slider min/max + HISTOGRAM signature pattern + 2 input "€50" / "€2600+"', 'Identico desktop · scrollable se largo'),
    ('4. Stanze e letti', 'Stepper ± per Camere / Letti / Bagni', 'Identico desktop'),
    ('5. Servizi', 'Checkbox grid (WiFi / AC / Cucina / Lavatrice...) con "Mostra di più"', 'Identico desktop'),
    ('6. Opzioni di prenotazione', '4 chip toggle (Prenotazione immediata / Self check-in / Cancellazione gratuita / Animali ammessi)', 'Identico desktop'),
    ('7. Alloggi straordinari', '2 tile (Amato dagli ospiti / Luxe)', 'Identico desktop'),
    ('8-10. Tipo accordion / Accessibilità / Lingua', 'Accordion ▼ collapsed default', 'Identico desktop'),
    ('Footer fisso', '"Cancella tutto" sx + "Mostra oltre 1.000 alloggi" CTA NERO #222222 dx (LIVE COUNT update ogni filtro)', '⚠️ CTA NERO mobile vs ROSA #FF385C desktop! Pattern intenzionale.'),
]

ws.cell(row=fr+2, column=1, value='Sezione').font = FONT_HEAD_TBL
ws.cell(row=fr+2, column=1).fill = PatternFill('solid', start_color='374151', end_color='374151')
ws.cell(row=fr+2, column=2, value='Pattern UI').font = FONT_HEAD_TBL
ws.cell(row=fr+2, column=2).fill = PatternFill('solid', start_color='374151', end_color='374151')
ws.cell(row=fr+2, column=3, value='Note vs desktop').font = FONT_HEAD_TBL
ws.cell(row=fr+2, column=3).fill = PatternFill('solid', start_color='374151', end_color='374151')
for c in range(1, 4):
    ws.cell(row=fr+2, column=c).alignment = ALIGN_C
    ws.cell(row=fr+2, column=c).border = BOX_BORDER

for r_idx, row in enumerate(filter_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=fr + 3 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L
        cell.border = BOX_BORDER

# Sezione Drag Handle
dr = fr + 3 + len(filter_rows) + 2
ws.cell(row=dr, column=1, value='DRAG HANDLE — Map ↔ Drawer toggle (pattern bottom-sheet)').font = Font(name='Arial', size=11, color='000000', bold=True)
ws.cell(row=dr, column=1).fill = PatternFill('solid', start_color='fce7f3', end_color='fce7f3')
ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=3)

ws.cell(row=dr+1, column=1, value='Linea grigia centrata orizzontalmente · 157×0px coordinate (visualmente pill 30×4) · sotto la mappa, sopra al drawer').font = FONT_LBL
ws.merge_cells(start_row=dr+1, start_column=1, end_row=dr+1, end_column=3)

handle_rows = [
    ('Stato 1: Drawer chiuso (default)', 'Map fullscreen-top 363×788 · drawer peek-bottom (mostra h1 "Oltre 1.000 alloggi" + prima card)'),
    ('Stato 2: Drawer aperto (drag-up)', 'Map collassa a peek-top · drawer espanso fullscreen mostra lista cards (1-col)'),
    ('Trigger', 'Drag verticale sul handle (touch swipe up/down) — NO tap-toggle (handle è decorativo, non button)'),
    ('URL state', 'drawer_open=true|false in querystring · permette deep-link in stato esplicito'),
    ('UX rationale', 'L\'utente sceglie quanto della mappa vs lista vedere. Default: mappa per esplorazione geografica, drawer drag per leggere lista.'),
]
for r_idx, row in enumerate(handle_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=dr + 2 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L
        cell.border = BOX_BORDER

# Note finali con confronto LivingApple
nr = dr + 2 + len(handle_rows) + 2
ws.cell(row=nr, column=1, value='IMPLICAZIONI per redesign LivingApple wizardstep1:').font = Font(name='Arial', size=10, bold=True)
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=3)

implications = [
    '1. La pillola searchbar mobile è UN SOLO touch target (vs 3 segmenti desktop) → in LivingApple già funziona (HomeSearch è bottone unico).',
    '2. La modale searchbar è ACCORDION step-by-step (Dove → Quando → Chi) → in LivingApple wizardstep1 ha già una flow simile (residenza → date → ospiti → offerta).',
    '3. La separazione Searchbar (modificare ricerca) vs Filtri (raffinare) è netta in Airbnb → utile mantenere in LivingApple: chip-bar above-fold per filtri rapidi + modale Filtri completa.',
    '4. CTA mobile NERO vs ROSA desktop → LivingApple potrebbe usare pattern simile (grigio scuro su mobile per non saturare, brand color su desktop dove c\'è più white-space).',
    '5. LIVE COUNT footer "Mostra oltre 1.000 alloggi" → LivingApple non ha questo, da introdurre nei filtri.',
    '6. Drag handle map↔list → NON applicabile a LivingApple (non abbiamo mappa, non è il nostro fulcro UX).',
    '7. Bottom nav 5 voci → NON applicabile a LivingApple (non è una platform multipage tipo airbnb/booking).',
]
for i, n in enumerate(implications):
    cell = ws.cell(row=nr + 1 + i, column=1, value=n)
    cell.font = FONT_LBL
    cell.alignment = ALIGN_L
    ws.merge_cells(start_row=nr + 1 + i, start_column=1, end_row=nr + 1 + i, end_column=3)


# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
wb.save(out_path)
print(f'OK: {out_path} ({os.path.getsize(out_path)} bytes) · {len(wb.sheetnames)} fogli totali')
print(f'Fogli: {wb.sheetnames}')
