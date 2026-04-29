"""
Genera docs/ux/bookingsidebar-diff-vs-mockup.xlsx — tabella diff fra il
mockup utente (docs/ux/mio desiderio.xlsx foglio wizardstep1, parte sidebar dx)
e lo stato attuale di components/wizard/BookingSidebar.tsx step=1.

Sessione 2026-04-29.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

C_HEADER_BG = '374151'
C_OK = 'd1fae5'      # verde chiaro = già OK
C_FIX = 'fef3c7'     # giallo = da modificare
C_REMOVE = 'fee2e2'  # rosso chiaro = mancante (rimuovere?)
C_BORDER = '9ca3af'

FONT_LBL = Font(name='Arial', size=10, color='1f2937')
FONT_LBL_BOLD = Font(name='Arial', size=10, color='1f2937', bold=True)
FONT_HEAD = Font(name='Arial', size=11, color='ffffff', bold=True)
FONT_TITLE = Font(name='Arial', size=12, color='000000', bold=True)

THIN = Side(style='thin', color=C_BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)


def make_header(ws, row, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = FONT_HEAD
        c.fill = PatternFill('solid', start_color=C_HEADER_BG, end_color=C_HEADER_BG)
        c.alignment = ALIGN_C
        c.border = BOX


wb = Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────────────────────────────────────
# FOGLIO 1: diff
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('diff')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 55
ws.column_dimensions['E'].width = 14

ws['A1'] = 'BookingSidebar.tsx (step=1) — diff fra mockup utente e codice attuale'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:E1')

ws['A2'] = 'Riferimenti: docs/ux/mio desiderio.xlsx (foglio wizardstep1, colonne CC-DF) · components/wizard/BookingSidebar.tsx step=1'
ws['A2'].font = Font(name='Arial', size=9, color='6b7280', italic=True)
ws.merge_cells('A2:E2')

make_header(ws, 4, ['#', 'Mockup utente (riga sidebar dx)', 'Attuale BookingSidebar.tsx step=1', 'Cosa fare', 'Stato'])

rows = [
    (1, '"Seleziona un appartamento dalla lista" italic grigio in cima',
     'C\'è t.selectRoomMsg ma SOTTO la foto, in posizione sbagliata',
     'Spostare in cima, prima della foto', 'fix'),
    (2, 'Foto hero 332×160 (.booking-sidebar__hero-img cover)',
     'Foto hero da Cloudinary (fallback _DSC2502_laqzeh.jpg)',
     '— (già OK)', 'ok'),
    (3, 'Nome casa (.section-title-secondary)',
     'Nome casa (.section-title-secondary)',
     '— (già OK)', 'ok'),
    (4, '"CARATTERISTICHE" label uppercase (.label-uppercase-muted)',
     'C\'è t.propertySection con .label-uppercase-muted',
     '— (già OK)', 'ok'),
    (5, '4 camere · 3 bagni · 12 ospiti · Piscina · Giardino (lista features)',
     'C\'è .booking-sidebar__feature-list con stesso contenuto',
     'Migrare alla classe master .feature-list (uniforme con WizardStep1 e residenze)', 'fix'),
    (6, 'Banner ⚡ Consumi energetici (.banner.banner--info) DOPO caratteristiche',
     'Banner Consumi PRIMA delle caratteristiche (ordine invertito)',
     'Spostare dopo le caratteristiche', 'fix'),
    (7, 'Banner 🔐 Deposito cauzionale — €(variabile) (.banner.banner--warning) subito dopo Consumi',
     'Banner Deposito DOPO il prezzo, non dopo Consumi (posizione lontana)',
     'Spostare subito dopo Consumi (banner adiacenti)', 'fix'),
    (8, 'Date · Modifica btn 81×32',
     'Date c\'è · Modifica btn appare SOLO in step=2 (callback onEditDates undefined in step1)',
     'Aggiungere Modifica anche in step=1 (con callback "modifica date")', 'fix'),
    (9, 'Ospiti · Modifica btn 81×32',
     'Ospiti c\'è · Modifica btn idem (solo step=2)',
     'Aggiungere Modifica anche in step=1 (con callback "modifica ospiti")', 'fix'),
    (10, 'Dettagli del prezzo (h2)',
     'Sezione c\'è con .label-uppercase-muted',
     '— (già OK)', 'ok'),
    (11, '4 notti × per-night · strike (661,28€) + nuovo (351,48€) — price drop',
     'Mostra prezzo notte+totale, NO strike+nuovo (solo strike se voucher applicato in step2)',
     'Aggiungere strike+nuovo quando c\'è price drop (sconto host)', 'fix'),
    (12, 'Imposta di soggiorno · valore',
     'C\'è (touristTax con formatTouristTaxNote)',
     '— (già OK)', 'ok'),
    (13, 'Totale EUR · valore (.booking-sidebar__total-value blu primary)',
     'C\'è (color: var(--color-primary))',
     '— (già OK)', 'ok'),
    (14, '"Riepilogo dei costi" link → apre modale 568×309 con breakdown dettagliato',
     'NON c\'è',
     'Aggiungere link che apre modale stile Airbnb (4 voci: soggiorno + servizio + tasse + totale)', 'fix'),
    (15, 'CTA "Continua" .booking-sidebar__cta .btn.btn--primary',
     'C\'è (showContinua basato su selectedOfferId)',
     '— (già OK)', 'ok'),
    ('M1', '(niente sezione "Cancellazione" nel mockup)',
     'Esiste sezione "CANCELLAZIONE" con .label-uppercase-muted + testo policy (offerName — condizione)',
     'DOMANDA: rimuovere del tutto o mantenere?', 'remove'),
    ('M2', '(niente footer CIN/CIR nel mockup)',
     'Esiste .booking-sidebar__footer con "CIN: " + numero CIN',
     'DOMANDA: rimuovere o mantenere (obbligo legale IT)?', 'remove'),
]

for r_idx, row in enumerate(rows):
    for c_idx, val in enumerate(row[:4]):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx in (0, 3) else FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (1, 2, 3) else ALIGN_C
        cell.border = BOX

    # Stato column (E)
    state = row[4]
    label = '✓ già OK' if state == 'ok' else ('🔧 da fixare' if state == 'fix' else '❓ mancante')
    color = C_OK if state == 'ok' else (C_FIX if state == 'fix' else C_REMOVE)
    cell = ws.cell(row=5 + r_idx, column=5, value=label)
    cell.font = FONT_LBL_BOLD
    cell.alignment = ALIGN_C
    cell.border = BOX
    cell.fill = PatternFill('solid', start_color=color, end_color=color)


# ─────────────────────────────────────────────────────────────────────────────
# FOGLIO 2: domande aperte
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.create_sheet('domande')
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 14

ws['A1'] = 'Domande aperte prima di toccare BookingSidebar.tsx'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:D1')

make_header(ws, 3, ['#', 'Domanda', 'Opzioni', 'Risposta utente'])

questions = [
    ('1', 'Cancellazione policy (sezione "CANCELLAZIONE" attuale con testo "Non rimborsabile — Pagamento entro 48h" ecc.)',
     'A) Rimuovere del tutto · B) Mantenere · C) Spostare altrove'),
    ('2', 'Footer CIN/CIR ("CIN: " obbligo legale IT per locazioni turistiche)',
     'A) Rimuovere · B) Mantenere ma più piccolo · C) Mantenere così'),
    ('3', '"Riepilogo dei costi" link — apre cosa?',
     'A) Modale stile Airbnb /book con 4 voci (Soggiorno + Servizio + Tasse + Totale) · B) Solo decorativo (no apertura)'),
    ('4', '"Modifica" btn in step=1 — cosa fa al click?',
     'A) Date → torna a HomeSearch su / · B) Date → modale date picker inline · C) Ospiti → modale stepper · D) Tutti i Modifica → step indietro a HomeSearch'),
    ('5', '"Strike + nuovo" prezzo (es. 661,28€ → 351,48€)',
     'A) Solo se c\'è price drop reale (offer.discountedPrice presente) · B) Sempre con tariffa media 60 notti come fa Airbnb (servirebbe API che ora non c\'è)'),
    ('6', 'Ordine sezioni: Caratteristiche → Banner Consumi → Banner Deposito → Date+Modifica → Ospiti+Modifica → Prezzo',
     'Confermi ordine?'),
]

for r_idx, row in enumerate(questions):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (1, 2) else ALIGN_C
        cell.border = BOX
    # Empty answer cell
    cell = ws.cell(row=4 + r_idx, column=4, value='')
    cell.border = BOX
    cell.fill = PatternFill('solid', start_color='f9fafb', end_color='f9fafb')


# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'docs', 'ux', 'bookingsidebar-diff-vs-mockup.xlsx'))
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f'OK: {out_path} ({os.path.getsize(out_path)} bytes) · {len(wb.sheetnames)} fogli: {wb.sheetnames}')
