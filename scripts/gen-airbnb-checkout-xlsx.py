"""
Genera docs/ux/airbnb-checkout-analysis.xlsx — analisi pagina checkout (Invia richiesta) Airbnb.
Dati estratti via Chrome MCP (Claude in Chrome) su www.airbnb.it/book/stays/...
viewport 2752×927 (XL desktop) + DevTools responsive 1920/1280/1024/768/480.
Listing: Pellini 3 - 180 mq · 4-8 mag 2026 · 1 adulto · Giuseppe host · request-to-book.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

# Palette (coerente con airbnb-pattern-analysis.xlsx)
C_HEADER = 'd1fae5'
C_TITLE  = 'fef3c7'
C_FORM   = 'dbeafe'
C_SUMMARY = 'fce7f3'
C_GAP    = 'f3f4f6'
C_PHOTO  = 'd1d5db'
C_CTA_ROSA  = 'FF385C'
C_CTA_NERO  = '222222'
C_BORDER = '9ca3af'
C_WHITE  = 'ffffff'
C_DIVIDER = 'e5e7eb'
C_RADIO_SEL = '222222'

FONT_LBL = Font(name='Arial', size=8, color='1f2937')
FONT_LBL_BOLD = Font(name='Arial', size=8, color='1f2937', bold=True)
FONT_LBL_SMALL = Font(name='Arial', size=7, color='6b7280', italic=True)
FONT_TITLE = Font(name='Arial', size=11, color='000000', bold=True)
FONT_CTA_W = Font(name='Arial', size=10, color='ffffff', bold=True)
FONT_HEAD_TBL = Font(name='Arial', size=10, color='ffffff', bold=True)

ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN = Side(style='thin', color=C_BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_grid(ws, n_cols=110, n_rows=110, col_w=2, row_h=15):
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(1, n_rows + 1):
        ws.row_dimensions[r].height = row_h


def draw_block(ws, range_str, fill_color, label='', font=None, merge=True, border=True):
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for c in row:
            c.fill = fill
            if border:
                c.border = BOX
    if merge:
        ws.merge_cells(range_str)
        cell = ws[f'{get_column_letter(min_col)}{min_row}']
        if label:
            cell.value = label
        cell.font = font if font else FONT_LBL_BOLD
        cell.alignment = ALIGN_C


def draw_container(ws, range_str, fill_color=None, border=True):
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    fill = PatternFill('solid', start_color=fill_color, end_color=fill_color) if fill_color else None
    for ri, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        for ci, c in enumerate(row):
            if fill:
                c.fill = fill
            if border:
                left = THIN if ci == 0 else None
                right = THIN if ci == max_col - min_col else None
                top = THIN if ri == 0 else None
                bottom = THIN if ri == max_row - min_row else None
                c.border = Border(left=left, right=right, top=top, bottom=bottom)


def write_label(ws, addr, text, font=None, align=None):
    ws[addr] = text
    ws[addr].font = font if font else FONT_LBL
    if align:
        ws[addr].alignment = align


def make_table_header(ws, row, headers, color='374151'):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = FONT_HEAD_TBL
        c.fill = PatternFill('solid', start_color=color, end_color=color)
        c.alignment = ALIGN_C
        c.border = BOX


# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 1: checkout-desktop-grid (sketch visuale layout 2-col)
# ═════════════════════════════════════════════════════════════════════════════
# Viewport 1920px → 110 col × ~17.5px ≈ 1920px
ws = wb.create_sheet('checkout-desktop-grid')
set_grid(ws, n_cols=110, n_rows=110)

ws['A1'] = 'Airbnb desktop CHECKOUT — Invia una richiesta di prenotazione (viewport 1920×943 misurato + 2752×927 ultrawide)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:DF1')

# Header (logo top-left)
draw_container(ws, 'A3:DF6', C_HEADER, border=True)
draw_block(ws, 'B3:F5', C_WHITE, '🏠 airbnb', font=FONT_LBL_BOLD, border=True)
write_label(ws, 'B6', 'header  102×80 logo top-left  ·  no nav search ·  no avatar/menu', FONT_LBL_SMALL)
ws.merge_cells('B6:DF6')

# H1 risultati (x=455 → col ~26, w=1010)
write_label(ws, 'AA8', '←  Invia una richiesta di prenotazione  (h1 1010×36 a y=112)  · back btn 16×16', FONT_LBL_BOLD)
ws.merge_cells('AA8:CC8')

# === FORM COL (azzurro) cols AA:BS (col 27-71, ~45 col ≈ 540px form) ===
draw_block(ws, 'AA10:BS10', C_FORM,
           '↓ FORM COLUMN  x=455  w=540  h=1091  ·  scroll-h totale 1396 ·  margine sx 455 / dx 480',
           font=FONT_LBL_BOLD)

# Card 1: Scegli quando pagare (188h → 11 righe a y=172)
draw_container(ws, 'AA12:BS22', C_WHITE, border=True)
write_label(ws, 'AB13', 'CARD 1 — Scegli quando pagare  540×188', FONT_LBL_BOLD)
ws.merge_cells('AB13:BS13')
write_label(ws, 'AB15', 'Paga 389,48 € subito  ●  (selected)', FONT_LBL)
ws.merge_cells('AB15:BS15')
write_label(ws, 'AB17', 'Paga in 3 rate con Klarna  ○', FONT_LBL)
ws.merge_cells('AB17:BS17')
write_label(ws, 'AB18', 'Paga in 3 rate da 129,82 € senza interessi · Ulteriori informazioni', FONT_LBL_SMALL)
ws.merge_cells('AB18:BS18')

# Card 2: Metodo di pagamento (102h → 6 righe)
draw_container(ws, 'AA24:BS30', C_WHITE, border=True)
write_label(ws, 'AB25', 'CARD 2 — Metodo di pagamento  540×102', FONT_LBL_BOLD)
ws.merge_cells('AB25:BS25')
write_label(ws, 'AB27', '🏦 VISA  4548  (saved)', FONT_LBL)
ws.merge_cells('AB27:BJ27')
draw_block(ws, 'BK27:BS27', 'f3f4f6', 'Modifica  81×32', font=FONT_LBL, border=True)
write_label(ws, 'AB29', 'logo row: VISA · MC · AMEX · PayPal · GPay  (accepted brands 13×8 each)', FONT_LBL_SMALL)
ws.merge_cells('AB29:BS29')

# Card 3: Scrivi un messaggio (316h → 19 righe)
draw_container(ws, 'AA32:BS50', C_WHITE, border=True)
write_label(ws, 'AB33', 'CARD 3 — Scrivi un messaggio all\'host  540×316', FONT_LBL_BOLD)
ws.merge_cells('AB33:BS33')
write_label(ws, 'AB35', 'Prima di continuare, racconta qualcosa sul tuo viaggio a Giuseppe e spiega perché questo alloggio fa al caso tuo.', FONT_LBL_SMALL)
ws.merge_cells('AB35:BS35')
draw_block(ws, 'AB37:AC38', C_PHOTO, '👤', font=FONT_LBL_SMALL, border=True)
write_label(ws, 'AE37', 'Giuseppe', FONT_LBL_BOLD)
write_label(ws, 'AE38', 'Host dal 2026', FONT_LBL_SMALL)
draw_container(ws, 'AB40:BS49', 'f9fafb', border=True)
write_label(ws, 'AC42', 'Esempio: "Ciao Giuseppe, io e il mio partner stiamo andando al matrimonio di un\'amica..."', FONT_LBL_SMALL)
ws.merge_cells('AC42:BR42')

# Card 4: Vuoi assicurazione (204h → 12 righe)
draw_container(ws, 'AA52:BS63', C_WHITE, border=True)
write_label(ws, 'AB53', 'CARD 4 — Vuoi aggiungere un\'assicurazione di viaggio?  540×204  (UPSELL)', FONT_LBL_BOLD)
ws.merge_cells('AB53:BS53')
write_label(ws, 'AB55', 'Sì, aggiungi per 17,38 €', FONT_LBL_BOLD)
ws.merge_cells('AB55:BJ55')
draw_block(ws, 'BK55:BS55', 'f3f4f6', 'Aggiungi  82×32', font=FONT_LBL, border=True)
write_label(ws, 'AB56', 'Disponibile solo al momento della prenotazione.', FONT_LBL_SMALL)
ws.merge_cells('AB56:BS56')
write_label(ws, 'AB59', 'Ottieni un rimborso fino al 100% del costo del tuo soggiorno se effettui una cancellazione', FONT_LBL_SMALL)
ws.merge_cells('AB59:BS59')
write_label(ws, 'AB60', 'per motivi coperti. Inoltre, godi della copertura su voli e attività.  ·  Cosa copre', FONT_LBL_SMALL)
ws.merge_cells('AB60:BS60')

# Disclaimer
write_label(ws, 'AA66', 'L\'host ha 24 ore per confermare la prenotazione, che ti verrà addebitata solo dopo che avrà accettato la tua richiesta.', FONT_LBL_SMALL)
ws.merge_cells('AA66:BS66')
write_label(ws, 'AA68', 'Cliccando sul pulsante, accetto i  termini di prenotazione.', FONT_LBL_SMALL)
ws.merge_cells('AA68:BS68')

# CTA finale
draw_block(ws, 'AA70:BS72', C_CTA_ROSA,
           'Invia una richiesta di prenotazione  ·  CTA 540×48 ROSA #FF385C  border-radius 12  padding 14×24',
           font=FONT_CTA_W)

# === GAP fra colonne (105 px = 8 col) ===
draw_container(ws, 'BT12:CB72', C_GAP, border=False)
write_label(ws, 'BT35', 'gap\n105 px', Font(name='Arial', size=7, color='6b7280'),
            align=Alignment(horizontal='center', vertical='center', wrap_text=True))
ws.merge_cells('BT35:CB35')

# === SUMMARY COL (rosa) cols CC:DF (col 81-110, ~30 col ≈ 340px) ===
draw_block(ws, 'CC10:DF10', C_SUMMARY,
           '↓ SUMMARY COLUMN  x=1100  w=340  position:STICKY top:32px  ·  altezza 544px',
           font=FONT_LBL_BOLD)

# Hero section: foto + titolo
draw_container(ws, 'CC12:DF22', C_WHITE, border=True)
draw_block(ws, 'CD13:CJ19', C_PHOTO, '[FOTO\n103×103]', font=FONT_LBL, border=True)
write_label(ws, 'CL14', 'Pellini 3 - 180 mq', FONT_LBL_BOLD)
ws.merge_cells('CL14:DE14')
write_label(ws, 'CL15', '(h2 221×24)', FONT_LBL_SMALL)
ws.merge_cells('CL15:DE15')
write_label(ws, 'CC21', 'Questa prenotazione non è rimborsabile.  ·  Modifica termini', FONT_LBL_SMALL)
ws.merge_cells('CC21:DF21')

# Date + Modifica
draw_container(ws, 'CC24:DF28', C_WHITE, border=True)
write_label(ws, 'CD25', 'Date', FONT_LBL_BOLD)
write_label(ws, 'CD26', '04–08 mag 2026', FONT_LBL)
draw_block(ws, 'DA25:DE26', 'f3f4f6', 'Modifica  81×32', font=FONT_LBL, border=True)

# Ospiti + Modifica
draw_container(ws, 'CC30:DF34', C_WHITE, border=True)
write_label(ws, 'CD31', 'Ospiti', FONT_LBL_BOLD)
write_label(ws, 'CD32', '1 adulto', FONT_LBL)
draw_block(ws, 'DA31:DE32', 'f3f4f6', 'Modifica  81×32', font=FONT_LBL, border=True)

# Dettagli del prezzo
draw_container(ws, 'CC36:DF50', C_WHITE, border=True)
write_label(ws, 'CD37', 'Dettagli del prezzo  (h2)', FONT_LBL_BOLD)
ws.merge_cells('CD37:DE37')
write_label(ws, 'CD39', '4 notti a 87,87 €', FONT_LBL)
write_label(ws, 'DC39', '661,28€  351,48€', FONT_LBL)
write_label(ws, 'CD40', '(strike + nuovo · price drop)', FONT_LBL_SMALL)
ws.merge_cells('CD40:DE40')
write_label(ws, 'CD42', 'Tasse', FONT_LBL)
write_label(ws, 'DC42', '38,00 €', FONT_LBL)
write_label(ws, 'CD45', 'Totale  EUR', FONT_LBL_BOLD)
write_label(ws, 'DC45', '389,48 €', FONT_LBL_BOLD)
write_label(ws, 'CD48', 'Riepilogo dei costi  (link → modale 568×309)', FONT_LBL_SMALL)
ws.merge_cells('CD48:DE48')

# STICKY indicator
write_label(ws, 'CC55', '◄ STICKY — resta in vista mentre form scrolla ►', FONT_LBL_BOLD,
            align=Alignment(horizontal='center'))
ws.merge_cells('CC55:DF55')

# Note totali
write_label(ws, 'A75', 'CALCOLO LAYOUT (a 1920):', FONT_LBL_BOLD)
notes = [
    '• viewport 1920 → margine sx 455 + FORM 540 + gap 105 + SUMMARY 340 + margine dx 480 = 1920',
    '• Container centrale CHECKOUT: 985 px wide (form 540 + gap 105 + summary 340) — più stretto del search results (1526)',
    '• Form col 540 wide · 5 cards a gap 24px tra una e l\'altra (Scegli quando · Metodo · Messaggio · Assicurazione · Disclaimer + CTA)',
    '• Summary col 340 wide · STICKY top:32 · contiene: hero(foto+titolo) + Date + Ospiti + Dettagli prezzo + link Riepilogo costi',
    '• CTA finale ROSA #FF385C border-radius 12 (vs filtri/Modifica btn neutri grigi)',
    '• Modifica btn standard: 81×32 (Date · Ospiti · Metodo pagamento — tutti uguali, design-system primary)',
    '• A 2752 (XL ultrawide): stesso container 985, summary diventa 390w sticky (più largo)',
    '• Il summary è sticky SOLO a 2752+ · da 1920 in giù sumPos:static (parent css media-query)',
]
for i, n in enumerate(notes):
    cell = ws.cell(row=77 + i, column=1, value=n)
    cell.font = FONT_LBL
    ws.merge_cells(start_row=77 + i, start_column=1, end_row=77 + i, end_column=110)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 2: checkout-desktop-data
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('checkout-desktop-data')
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 60

ws['A1'] = 'Airbnb desktop CHECKOUT — tabella dimensioni misurate via Chrome MCP (1920×943)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:H1')

ws['A2'] = 'URL: airbnb.it/book/stays/... · Pellini 3 - 180 mq · 04-08 mag 2026 · 1 adulto · request-to-book · scroll-h 1396'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:H2')

make_table_header(ws, 4, ['Elemento', 'Selettore / aria', 'x px', 'y px', 'w px', 'h px', 'parent', 'note'])

data_rows = [
    ('viewport', 'window', 0, 0, 1920, 943, 'browser', 'misurato in DevTools responsive · scroll-h 1396'),
    ('header logo', 'a[href="/"]', 24, 0, 102, 80, 'banner', 'logo airbnb full SVG · top-left · NO nav inline'),
    ('back arrow', 'main button (icon)', 410, 126, 16, 16, 'main', '← icona freccia · ritorna alla scheda alloggio'),
    ('h1 page title', 'h1', 455, 112, 1010, 36, 'main', '"Invia una richiesta di prenotazione" (request-to-book; "Conferma e paga" per instant-book)'),
    ('FORM column container', 'div.cgwpmcp', 455, 172, 540, 1091, 'main', 'colonna SX · contiene 5 cards · gap 24px fra cards'),
    ('  card1 Scegli quando pagare', '(div card)', 455, 172, 540, 188, 'form-col', 'radio "Paga subito" + "Paga in 3 rate Klarna"'),
    ('    radio Paga subito €', 'input[type=radio] (selected)', 1357, 237, 22, 22, 'card1', 'CHECKED · radio circle a destra del label'),
    ('    radio Paga 3 rate Klarna', 'input[type=radio]', 1357, 294, 22, 22, 'card1', 'unchecked · BNPL provider integrato'),
    ('  card2 Metodo di pagamento', '(div card)', 455, 384, 540, 122, 'form-col', 'mostra carta saved + brand row + Modifica'),
    ('    VISA logo (saved)', 'img[alt="VISA"]', 480, 441, 20, 20, 'card2', 'logo carta selezionata'),
    ('    Modifica btn', 'button "Modifica"', 890, 419, 81, 32, 'card2', 'apre modale Metodo pagamento (568×800)'),
    ('    brand row (accept)', 'img[alt="MC"|"AMEX"|"PayPal"|"GPay"]', 461, 498, 'vari', 8, 'card2', 'mini-icone metodi accettati (13×8 each)'),
    ('  card3 Scrivi messaggio', '(div card)', 455, 530, 540, 316, 'form-col', 'textarea grande · Giuseppe avatar + label'),
    ('    avatar host', 'img round 32x32', '~480', 555, 32, 32, 'card3', 'foto host circle · "Giuseppe" / "Host dal 2026"'),
    ('    textarea', 'textarea', '~480', 615, '~500', '~210', 'card3', 'placeholder lungo "Esempio: Ciao Giuseppe..."'),
    ('  card4 Assicurazione (UPSELL)', '(div card)', 455, 870, 540, 204, 'form-col', '"Sì, aggiungi per 17,38 €" + Aggiungi btn + sub-text'),
    ('    Aggiungi btn (insurance)', 'button[aria-label="Aggiungi assicurazione"]', 888, 922, 82, 32, 'card4', 'apre sub-modal Termini assicurazione'),
    ('  disclaimer 24h', '(div text)', 455, 1098, 540, 50, 'form-col', '"L\'host ha 24 ore per confermare..."'),
    ('  disclaimer terms', '(div text)', 455, 1183, 540, 18, 'form-col', '"Cliccando sul pulsante, accetto i termini di prenotazione."'),
    ('  CTA Invia richiesta ★', 'button (final submit)', 455, 1215, 540, 48, 'form-col', 'ROSA #FF385C · border-radius 12 · padding 14×24 · font-weight 500 size 16'),
    ('SUMMARY column container', 'div.cperwgn', 1100, 197, 340, 'auto', 'main', 'colonna DX · STICKY top:32 (a 2752+) · static (a 1920-)'),
    ('  hero photo', 'img (listing thumb)', 1100, 197, 103, 103, 'summary', 'aspect 1:1 · square thumbnail dell\'alloggio'),
    ('  hero title', 'h2 "Pellini 3 - 180 mq"', 1219, 235, 221, 24, 'summary', 'h2 a destra della foto · dx-aligned'),
    ('  cancellation policy', '(div text)', 1100, 313, 340, 18, 'summary', '"Questa prenotazione non è rimborsabile."'),
    ('  Termini completi link', '(button text)', 1100, 337, 111, 18, 'summary', 'apre modale termini'),
    ('  Date label', '(div)', 1100, 389, 260, 18, 'summary', '"Date"'),
    ('  Date value', '(div)', 1100, 414, 108, 18, 'summary', '"04–08 mag 2026"'),
    ('  Date Modifica btn', 'button', 1376, 397, 81, 32, 'summary', 'apre modale modifica date · stesso 81×32 design-system'),
    ('  Ospiti label', '(div)', 1100, 466, 260, 18, 'summary', '"Ospiti"'),
    ('  Ospiti value', '(div)', 1100, 491, 50, 18, 'summary', '"1 adulto"'),
    ('  Ospiti Modifica btn', 'button', 1376, 474, 81, 32, 'summary', 'apre modale ospiti'),
    ('  Dettagli del prezzo h2', 'h2', 1100, 542, 340, 18, 'summary', 'sezione price breakdown'),
    ('    4 notti label', '(div)', 1100, '~580', '~140', 21, 'summary', '"4 notti a 87,87 €"'),
    ('    price strike+new', '(div)', 1370, '~580', '~125', 21, 'summary', '"661,28 €" strike + "351,48 €" final · price drop visible'),
    ('    Tasse label', '(div)', 1100, '~605', 50, 18, 'summary', '"Tasse"'),
    ('    Tasse value', '(div)', '~1410', '~605', 30, 18, 'summary', '"38,00 €"'),
    ('    Totale EUR (bold)', '(div)', 1100, '~650', 70, 21, 'summary', '"Totale  EUR"'),
    ('    Totale value (bold)', '(div)', '~1390', '~650', 50, 21, 'summary', '"389,48 €"'),
    ('    Riepilogo dei costi link', 'button (modal trigger)', 1100, 673, 120, 18, 'summary', 'apre modale 568×309 con breakdown COMPLETO (4 voci: soggiorno + servizio + tasse)'),
    ('footer', 'footer', 0, 1327, 1920, 69, 'viewport', '"Privacy · Termini · Dettagli dell\'azienda"'),
]

for r_idx, row in enumerate(data_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6, 7) else ALIGN_C
        cell.border = BOX
        if isinstance(val, str) and not val.startswith('  ') and c_idx == 0:
            cell.font = FONT_LBL_BOLD

note_row = 5 + len(data_rows) + 2
notes_data = [
    'PATTERN UX CHIAVE (desktop checkout):',
    '• 2-col asimmetrica: form 540 (61%) + summary 340 (39%) · gap 105 · container 985 wide centered',
    '• Summary STICKY top:32 (su ultrawide) — pattern signature: utente vede prezzo+CTA del riepilogo mentre scorre form',
    '• 5 cards form gap 24px · ognuna con padding interno · border 1px lieve',
    '• Modifica btn DESIGN-SYSTEM: 81×32 grigio neutro con border-radius 8 · usato per Date/Ospiti/Metodo pagamento',
    '• CTA finale ROSA #FF385C 540×48 border-radius 12 padding 14×24 · solo per ACTION irreversibile (submit)',
    '• Brand row in Metodo pagamento: micro-icone 13×8 sotto la card (visual reassurance metodi accettati)',
    '• Price transparency: "661,28 € → 351,48 €" strike+nuovo direttamente nel summary (price drop visible)',
    '• "Riepilogo dei costi" link → modale separata con BREAKDOWN COMPLETO (4 voci: soggiorno €308 + servizio Airbnb €43,48 + tasse €38 = €389,48)',
    '• NESSUN campo discount code visibile · NESSUN "save card for next time" (la carta è già saved)',
    '• Trust signals: SOLO il "termini di prenotazione" link · NO Stripe/secure logo · brand fa da trust signal',
    '• Login wall: pagina checkout richiede login (non visibile in screenshot perché già loggato)',
    '• Disclaimer 24h è SPECIFICO per request-to-book (instant-book non ha 24h wait)',
]
for i, n in enumerate(notes_data):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if i == 0 else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=8)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 3: breakpoints
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('breakpoints')
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 60

ws['A1'] = 'Airbnb checkout — 6 width misurate via Chrome MCP + DevTools responsive'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:G1')

make_table_header(ws, 3, ['Width (px)', 'Layout', 'Form col', 'Summary col', 'Container', 'CTA pattern', 'Note'])

bp_rows = [
    ('2752', '2-col + STICKY', '540w x=863', '390w sticky top:32', '985', 'ROSA "Invia richiesta" 540×48 bottom-form',
     'XL ultrawide (DPR 1.25 Windows 4K) · margini 863sx/889dx'),
    ('1920', '2-col', '540w x=455', '340w static', '985', 'ROSA "Invia richiesta" 540×48',
     'Desktop FHD · margini 455sx/480dx · sumPos diventa STATIC sotto 2200px'),
    ('1280', '2-col', '440w x=196', '318w static', '863', 'ROSA "Invia richiesta" 440×48',
     'Laptop · form si stringe da 540→440 (-100px) · summary -22'),
    ('1024', '2-col', '386w x=121', '282w static', '757', 'ROSA "Invia richiesta" 386×48',
     'Tablet landscape · form 386 minimo desktop · ancora side-by-side'),
    ('768', '🎯 STACKED', '440w x=164', '440w x=164 (TOP)', '440', 'ROSA "Invia richiesta" 440×48',
     '🎯 Breakpoint critico ~770: passa a 1-col STACKED · summary IN ALTO sopra al form (sumY=280, formY=783) · logo si riduce a 30w iconetta'),
    ('480', '🎯 WIZARD MOBILE', '432w fullscreen', '(integrato in step 4)', '432', 'NERO "Avanti" step 1-3 + ROSA "Invia richiesta" step 4',
     '🎯 Breakpoint critico ~520: ENTRA IN WIZARD MOBILE 4-step (drawer fullscreen) · CTA cambia colore (nero→rosa) all\'ultimo step · bottom nav 5 voci sempre presente'),
]
for r_idx, row in enumerate(bp_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=4 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (1, 2, 3, 5, 6) else ALIGN_C
        cell.border = BOX
        if isinstance(val, str) and ('🎯' in val or 'WIZARD' in val or 'STACKED' in val or 'STICKY' in val):
            cell.fill = PatternFill('solid', start_color='fef3c7', end_color='fef3c7')

note_row = 4 + len(bp_rows) + 2
notes_bp = [
    'BREAKPOINT CRITICI:',
    '• ~2200 px : summary diventa STICKY (è static fino a 1920)',
    '• ~770 px : 2-col → STACKED (summary in alto, form sotto, entrambi 440w · logo si riduce a icona 30w)',
    '• ~520 px : STACKED → WIZARD MOBILE 4-step (drawer fullscreen, bottom nav 5 voci, step indicator)',
    '',
    '4 ZONE LAYOUT:',
    '• ULTRAWIDE (≥2200): 2-col simmetrico + summary sticky (utente vede prezzo+CTA fissi)',
    '• DESKTOP (770-2200): 2-col asimmetrico statico, container ~760-985 wide',
    '• TABLET PORTRAIT (520-770): 1-col stacked, summary IN ALTO (review-first pattern)',
    '• MOBILE (<520): WIZARD MULTI-STEP fullscreen (4 step "Verifica → Messaggio → Assicurazione → Submit")',
    '',
    'CONFRONTO con LivingApple /paga (PagaClient.tsx):',
    '• LivingApple ha 1 breakpoint (mobile <768) · Airbnb ne ha 4 (ultrawide/desktop/tablet/mobile-wizard)',
    '• LivingApple su mobile resta SINGLE PAGE (form sopra summary inline) · Airbnb va in WIZARD 4-step su <520',
    '• Pattern wizard mobile è una scelta UX importante: riduce cognitive load step-by-step (pro) ma aumenta tap-to-conversion (contro)',
    '• Sticky summary su ultrawide è UX premium · LivingApple da considerare su /paga desktop',
]
for i, n in enumerate(notes_bp):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if (i == 0 or i == 5 or i == 11) else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=7)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 4: payment-modal (modale Metodo di pagamento)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('payment-modal')
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 70

ws['A1'] = 'Airbnb modale Metodo di pagamento — 568×800 centered (border-radius 32, bg-white)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:G1')

ws['A2'] = 'Trigger: click "Modifica" su card "Metodo di pagamento" nel checkout · 5 metodi attivi + 1 scaduta · footer Annulla/Fatto'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:G2')

make_table_header(ws, 4, ['Elemento', 'Tag', 'x px', 'y px', 'w px', 'h px', 'note / contenuto'])

modal_rows = [
    ('MODAL container', 'role=dialog', 1092, 64, 568, 800, 'centered viewport · bg #fff · border-radius 32 · NO padding (children handle)'),
    ('  close X', 'button (top-right)', 1620, 88, 16, 16, 'icona × · border-radius 50% (cerchio) · color #222'),
    ('  h1 title', 'h1 "Metodo di pagamento"', '~1300', '~85', 'auto', 24, 'centered horizontal'),
    ('  ── SAVED CARD ──', '(section)', 1116, 200, 520, 60, 'sezione carte già salvate'),
    ('    radio VISA 4548 (selected)', 'input[type=radio] checked', 1614, 211, 22, 22, '✓ default/selected · pattern card-saved'),
    ('  ── SECTION "Oppure paga con" ──', '(label)', 1116, 283, '~150', 18, 'separator label · sezione metodi alternativi'),
    ('    radio Carta credito/debito', 'input[type=radio]', 1614, 348, 22, 22, 'apre form add new card · brand: VISA + MC + AMEX'),
    ('    radio Apple Pay', 'input[type=radio]', 1614, 420, 22, 22, 'pattern wallet (mobile-friendly · fingerprint/FaceID)'),
    ('    radio PayPal', 'input[type=radio]', 1614, 493, 22, 22, 'redirect a PayPal login flow'),
    ('    radio Google Pay', 'input[type=radio]', 1614, 566, 22, 22, 'pattern wallet (Android/Chrome)'),
    ('  ── SECTION "Non disponibili" ──', '(label)', 1116, '~660', '~150', 18, 'separator · sezione carte scadute/non utilizzabili'),
    ('    radio Carta debito 0853 (Scaduta)', 'input[type=radio] disabled', 1614, '~720', 22, 22, 'opacità ridotta · text-color grey · DISABILITATO ma visibile (UX trasparenza)'),
    ('  ── FOOTER ──', '(div fixed bottom)', 1092, 799, 568, 64, 'sticky footer · 2 bottoni'),
    ('    Annulla btn', 'button (secondary text)', 1102, 799, 84, 48, 'TRANSPARENT bg · color #222 · border-radius 12 · padding 14 · text-only secondario'),
    ('    Fatto btn ★', 'button (primary)', 1524, 799, 112, 48, '★ NERO #222 · color #fff · border-radius 12 · padding 14×24 · font-weight 500'),
]

for r_idx, row in enumerate(modal_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 6) else ALIGN_C
        cell.border = BOX
        if isinstance(val, str) and val.strip().startswith('  ──'):
            cell.font = FONT_LBL_BOLD
            cell.fill = PatternFill('solid', start_color='f3f4f6', end_color='f3f4f6')

note_row = 5 + len(modal_rows) + 2
notes_pm = [
    'PATTERN UX CHIAVE (modale Metodo pagamento):',
    '• 5 metodi attivi: Carta saved · Carta nuova · Apple Pay · PayPal · Google Pay  (NESSUN Klarna qui — quello è in card "Scegli quando pagare")',
    '• Carte SCADUTE mostrate ma DISABILITATE (UX trasparente: l\'utente vede la cronologia, non si chiede "dov\'è andata la mia carta?")',
    '• Sezione "Oppure paga con" è sempre presente anche con saved card (offre chiaro switch)',
    '• CTA modale "Fatto" è NERO #222 (NON rosa) — pattern intenzionale: il rosa è SOLO per il submit-finale-conversion del checkout',
    '• Annulla è transparent text-only (secondary action) · Fatto è solid primary (positive action)',
    '• border-radius 32px del modal · 12px del CTA · 50% del close X — gerarchia di rounded',
    '',
    'IMPLICAZIONI per LivingApple /paga:',
    '• LivingApple offre solo Stripe (carta) + PayPal · Aggiungere Apple Pay/Google Pay è banale via Stripe Elements (Wallet API)',
    '• La modale "Modifica metodo pagamento" è un pattern utile da introdurre se in futuro avremo carte salvate (Stripe Customer + saved methods)',
    '• Mostrare carte scadute con stato disabled è UX onesta · da imitare se gestiamo profili utente con cronologia',
]
for i, n in enumerate(notes_pm):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if (i == 0 or i == 7) else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=7)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 5: price-breakdown (modale Riepilogo dei costi)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('price-breakdown')
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 75

ws['A1'] = 'Airbnb modale Riepilogo dei costi — 568×309 centered (border-radius 32, bg-white)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:F1')

ws['A2'] = 'Trigger: click "Riepilogo dei costi" link nel summary card desktop · breakdown TRASPARENTE 4 voci'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:F2')

make_table_header(ws, 4, ['Voce', 'Tag', 'x px', 'y px', 'Valore', 'Note'])

pb_rows = [
    ('MODAL container', 'role=dialog', 1092, 309, 568, '309 (h)', 'bg #fff · border-radius 32'),
    ('  h1 title', 'h1', 1307, 331, 138, '"Riepilogo dei costi"', 'centered'),
    ('  close X', 'button (top-right)', '~1620', '~331', 16, '×', 'cerchio · close modale'),
    ('  ── BREAKDOWN ──', '(section)', 1116, 380, 520, '4 voci', 'list-style senza separatori grafici (solo spacing)'),
    ('  voce 1: Soggiorno', 'div (left)', 1116, 389, 'auto', '"4 notti · 04–08 mag"', 'label sx aligned'),
    ('    prezzo originale (strike)', 'span', 1507, 388, 61, '661,28 €', 'strike-through · grey'),
    ('    prezzo finale', 'span', 1572, 388, 64, '308,00 €', 'BLACK · final'),
    ('    sub-text price drop', '(div multi-line)', 1116, 417, 520, '"Di recente, Giuseppe ha abbassato il prezzo per queste date (rispetto alla tariffa media delle ultime 60 notti)."', '⭐ TRANSPARENCY · spiega WHY il prezzo è scontato (host ha abbassato)'),
    ('  voce 2: Costi servizio Airbnb', 'div (left)', 1116, '~470', 'auto', '"Costi del servizio Airbnb"', 'marketplace fee esplicito'),
    ('    valore servizio', 'span (right)', '~1580', '~470', '~50', '43,48 €', 'NON nascosto in totale (UX onesto)'),
    ('  voce 3: Tasse', 'div (left)', 1116, '~510', 'auto', 'Tasse', 'link "Tasse" è cliccabile (info)'),
    ('    valore tasse', 'span (right)', '~1580', '~510', '~50', '38,00 €', 'tasse cumulative'),
    ('  ── DIVIDER ──', '(border-top)', 1116, '~565', 520, 'auto', 'separator 1px'),
    ('  Totale EUR', 'div bold', 1116, '~595', 'auto', 'Totale EUR', 'bold · key value · cliccabile EUR (currency selector)'),
    ('    valore totale', 'span (right) bold', '~1580', '~595', '~60', '389,48 €', 'gross-everything · invariato'),
]

for r_idx, row in enumerate(pb_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL
        cell.alignment = ALIGN_L if c_idx in (0, 1, 4, 5) else ALIGN_C
        cell.border = BOX
        if isinstance(val, str) and val.strip().startswith('  ──'):
            cell.font = FONT_LBL_BOLD
            cell.fill = PatternFill('solid', start_color='f3f4f6', end_color='f3f4f6')

note_row = 5 + len(pb_rows) + 2
notes_pb = [
    'PATTERN UX CHIAVE (Riepilogo dei costi):',
    '• PRICE TRANSPARENCY è la firma di Airbnb su questo modale:',
    '  - mostra prezzo originale strike + prezzo finale lato a lato',
    '  - SPIEGA perché c\'è il discount ("Giuseppe ha abbassato il prezzo per queste date") — dato concreto, non marketing',
    '  - separa esplicitamente "Costi del servizio Airbnb" (marketplace fee) — non nascosto in "altre voci"',
    '  - tasse separate da soggiorno · tutto cumulato in "Totale EUR" finale',
    '• 4 voci totali: Soggiorno (con strike) + Servizio Airbnb + Tasse + Totale (no extra come cleaning, animal, ecc — non applicabili qui)',
    '• "EUR" è cliccabile (currency selector globale Airbnb) · Totale è bold · valori dx-aligned',
    '• Modale APPARE su desktop · su mobile è inline nello step 1 ("Dettagli" btn nel wizard apre modale)',
    '',
    'CONFRONTO con LivingApple /paga (BookingSidebar attuale):',
    '• LivingApple BookingSidebar mostra già breakdown: notti×prezzo + cleaning + servizio + tasse + totale (5 voci)',
    '• MA: LivingApple non SPIEGA il prezzo (es. "questo prezzo è scontato perché [reason]") — pattern Airbnb da imitare quando offer attive',
    '• LivingApple mostra "Costi servizio LivingApple" come fee — coerente con Airbnb pattern',
    '• Da introdurre: link "Riepilogo dettagliato" nella BookingSidebar per UX cleaner (compresso default + expandable per chi vuole dettagli)',
]
for i, n in enumerate(notes_pb):
    cell = ws.cell(row=note_row + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if (i == 0 or i == 9) else FONT_LBL
    ws.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=6)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 6: mobile-wizard-grid (sketch step 1 + step 4 affiancati)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('mobile-wizard-grid')
set_grid(ws, n_cols=70, n_rows=110, col_w=2, row_h=15)

ws['A1'] = 'Airbnb checkout mobile WIZARD — 480×840 viewport (DevTools responsive · sketch step 1 + step 4 affiancati)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:BR1')

# === STEP 1: Verifica e continua === (cols A:AF)
ws['A3'] = 'STEP 1 — Verifica e continua'
ws['A3'].font = Font(name='Arial', size=10, color='000000', bold=True)
ws['A3'].fill = PatternFill('solid', start_color='dbeafe', end_color='dbeafe')
ws.merge_cells('A3:AF3')

# Top: ← back / X close
draw_block(ws, 'B5:C6', C_WHITE, '←', font=FONT_LBL_BOLD, border=True)
draw_block(ws, 'AD5:AE6', C_WHITE, '×', font=FONT_LBL_BOLD, border=True)
write_label(ws, 'E5', 'Verifica e continua  (h1)', FONT_LBL_BOLD)
ws.merge_cells('E5:AB5')

# Card alloggio
draw_container(ws, 'B8:AF18', C_WHITE, border=True)
draw_block(ws, 'C9:G14', C_PHOTO, '[FOTO]', font=FONT_LBL_SMALL, border=True)
write_label(ws, 'I10', 'Pellini 3 - 180 mq', FONT_LBL_BOLD)
ws.merge_cells('I10:AE10')
write_label(ws, 'C16', 'Questa prenotazione non è rimborsabile.  ·  Modifica termini', FONT_LBL_SMALL)
ws.merge_cells('C16:AE16')

# Date + Modifica
write_label(ws, 'C20', 'Date', FONT_LBL_BOLD)
write_label(ws, 'C21', '04–08 mag 2026', FONT_LBL)
draw_block(ws, 'X20:AC21', 'f3f4f6', 'Modifica', font=FONT_LBL, border=True)

# Ospiti + Modifica
write_label(ws, 'C24', 'Ospiti', FONT_LBL_BOLD)
write_label(ws, 'C25', '1 adulto', FONT_LBL)
draw_block(ws, 'X24:AC25', 'f3f4f6', 'Modifica', font=FONT_LBL, border=True)

# Prezzo totale + Dettagli btn
write_label(ws, 'C28', 'Prezzo totale', FONT_LBL_BOLD)
write_label(ws, 'C29', '389,48 €, tasse incluse  EUR', FONT_LBL)
draw_block(ws, 'X28:AC29', 'f3f4f6', 'Dettagli', font=FONT_LBL, border=True)

# Section Scegli quando pagare
write_label(ws, 'C32', 'Scegli quando pagare  (h3)', FONT_LBL_BOLD)
ws.merge_cells('C32:AE32')
draw_container(ws, 'B34:AF42', C_WHITE, border=True)
write_label(ws, 'C35', 'Paga 389,48 € subito  ●', FONT_LBL_BOLD)
ws.merge_cells('C35:AE35')
write_label(ws, 'C38', 'Paga in 3 rate con Klarna  ○', FONT_LBL)
ws.merge_cells('C38:AE38')
write_label(ws, 'C39', 'Paga in 3 rate da 129,82 € senza interessi.  Ulteriori informazioni', FONT_LBL_SMALL)
ws.merge_cells('C39:AE39')

# Step indicator (4 dots)
draw_block(ws, 'C46:H46', C_CTA_NERO, '', merge=False, border=False)
draw_block(ws, 'J46:O46', C_DIVIDER, '', merge=False, border=False)
draw_block(ws, 'Q46:V46', C_DIVIDER, '', merge=False, border=False)
draw_block(ws, 'X46:AC46', C_DIVIDER, '', merge=False, border=False)
write_label(ws, 'C47', '── step indicator (4 dots, primo nero=current) ──', FONT_LBL_SMALL)
ws.merge_cells('C47:AE47')

# CTA Avanti NERO
draw_block(ws, 'B49:AF51', C_CTA_NERO, 'Avanti  (CTA NERO 432×40 br12)', font=FONT_CTA_W)

# Bottom nav (5 voci)
draw_container(ws, 'B53:AF55', C_WHITE, border=True)
write_label(ws, 'B53', '◄── BOTTOM NAV (5 voci sticky · Esplora · Preferiti · Viaggi · Messaggi · Profilo) ──►', FONT_LBL_BOLD,
            align=Alignment(horizontal='center'))
ws.merge_cells('B53:AF53')

# === STEP 4: Invia richiesta finale === (cols AH:BR)
ws['AH3'] = 'STEP 4 — Invia richiesta (review + submit)'
ws['AH3'].font = Font(name='Arial', size=10, color='000000', bold=True)
ws['AH3'].fill = PatternFill('solid', start_color='fce7f3', end_color='fce7f3')
ws.merge_cells('AH3:BR3')

# Top: × close (no back arrow al final step)
draw_block(ws, 'BP5:BQ6', C_WHITE, '×', font=FONT_LBL_BOLD, border=True)
write_label(ws, 'AI5', 'Invia una richiesta di prenotazione  (h1 multi-line)', FONT_LBL_BOLD)
ws.merge_cells('AI5:BL5')

# Card Metodo di pagamento (clickable → freccia)
draw_container(ws, 'AI8:BR12', C_WHITE, border=True)
write_label(ws, 'AJ9', 'Metodo di pagamento', FONT_LBL_BOLD)
write_label(ws, 'AJ10', '🏦 VISA  4548', FONT_LBL)
write_label(ws, 'BO9', '→', FONT_LBL_BOLD)
write_label(ws, 'AJ12', 'VISA · MC · AMEX · PayPal · GPay (brand row)', FONT_LBL_SMALL)

# Card Scrivi messaggio
draw_container(ws, 'AI14:BR18', C_WHITE, border=True)
write_label(ws, 'AJ15', 'Invia un messaggio all\'host', FONT_LBL_BOLD)
write_label(ws, 'AJ16', '"ciao ciao"  (preview)', FONT_LBL_SMALL)
write_label(ws, 'BO15', '→', FONT_LBL_BOLD)

# Card assicurazione (last chance upsell)
draw_container(ws, 'AI20:BR26', C_WHITE, border=True)
write_label(ws, 'AJ21', 'Vuoi aggiungere un\'assicurazione di viaggio?', FONT_LBL_BOLD)
ws.merge_cells('AJ21:BO21')
write_label(ws, 'AJ23', 'Sì, aggiungi per 17,38 €', FONT_LBL)
draw_block(ws, 'BK23:BP24', 'f3f4f6', 'Aggiungi', font=FONT_LBL, border=True)
write_label(ws, 'AJ25', '"Cosa copre"', FONT_LBL_SMALL)

# Dettagli prezzo inline
write_label(ws, 'AI28', 'Dettagli del prezzo  (h2)', FONT_LBL_BOLD)
ws.merge_cells('AI28:BR28')
write_label(ws, 'AI30', '4 notti a 87,87 €', FONT_LBL)
write_label(ws, 'BJ30', '661,28€  351,48€', FONT_LBL)
write_label(ws, 'AI32', 'Tasse', FONT_LBL)
write_label(ws, 'BJ32', '38,00 €', FONT_LBL)
write_label(ws, 'AI34', 'Totale  EUR', FONT_LBL_BOLD)
write_label(ws, 'BJ34', '389,48 €', FONT_LBL_BOLD)
write_label(ws, 'AI36', 'Riepilogo dei costi  (link)', FONT_LBL_SMALL)

# Disclaimer
write_label(ws, 'AI39', 'L\'host ha 24 ore per confermare...', FONT_LBL_SMALL)
ws.merge_cells('AI39:BR39')
write_label(ws, 'AI41', 'Cliccando, accetto i termini di prenotazione.', FONT_LBL_SMALL)
ws.merge_cells('AI41:BR41')

# CTA Invia richiesta ROSA (★ key change)
draw_block(ws, 'AH43:BR46', C_CTA_ROSA, 'Invia una richiesta di prenotazione  ·  ★ ROSA #FF385C 432×52', font=FONT_CTA_W)

# Footer mini
write_label(ws, 'AH48', 'Privacy · Termini · Dettagli dell\'azienda', FONT_LBL_SMALL,
            align=Alignment(horizontal='center'))
ws.merge_cells('AH48:BR48')

# Note finali
write_label(ws, 'A56', 'PATTERN UX CHIAVE (wizard mobile):', FONT_LBL_BOLD)
ws.merge_cells('A56:BR56')
notes_mw = [
    '• 4 STEP wizard fullscreen drawer · ognuno ha h1 diverso (Verifica → Scrivi messaggio → Vuoi assicurazione → Invia richiesta)',
    '• STEP indicator 4 dots in basso (primo nero = current, altri grigi)',
    '• CTA cambia COLORE all\'ultimo step: NERO "Avanti" (step 1-3) → ROSA "Invia richiesta" (step 4)',
    '• Step 4 ricapitola TUTTO: Metodo pagamento + Messaggio + Assicurazione (come 3 cards cliccabili → freccia per modificare) + Prezzo dettagliato',
    '• Step 3 (assicurazione) ha sub-modal CON termini insurance + 2 CTA "No, grazie" / "Conferma e aggiungi" — UX legal-compliance forte',
    '• Bottom nav 5 voci (Esplora/Preferiti/Viaggi/Messaggi/Profilo) RESTA presente DURANTE il wizard — si può uscire in qualsiasi momento',
    '• Top: ← back arrow + × close X (eccetto step 4 dove c\'è solo ×)',
    '• Pattern: progressive disclosure mobile (vs all-on-one-page desktop)',
]
for i, n in enumerate(notes_mw):
    cell = ws.cell(row=58 + i, column=1, value=n)
    cell.font = FONT_LBL
    ws.merge_cells(start_row=58 + i, start_column=1, end_row=58 + i, end_column=70)


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 7: mobile-wizard-flow (tabella step-by-step)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('mobile-wizard-flow')
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 50

ws['A1'] = 'Airbnb checkout mobile — Wizard 4-step flow (480×840 viewport)'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:E1')

ws['A2'] = 'Trigger entrata wizard: viewport <520 (rilevato da media query) · drawer fullscreen sostituisce la pagina checkout 2-col desktop'
ws['A2'].font = FONT_LBL_SMALL
ws.merge_cells('A2:E2')

make_table_header(ws, 4, ['Step (h1)', 'Cards / contenuto', 'Sub-elements', 'CTA / footer', 'Note UX'])

flow_rows = [
    ('1️⃣ Verifica e continua',
     'Card alloggio (foto + Pellini 3 - 180 mq) + Termini cancellazione',
     'Date 04-08 mag + Modifica · Ospiti 1 adulto + Modifica · Prezzo totale + "Dettagli" btn (apre Riepilogo modal) · "Modifica termini" link',
     'CTA NERO "Avanti" 432×40',
     'Step indicator dot 1/4 attivo · Sezione "Scegli quando pagare" (radio Paga subito/Klarna)'),
    ('2️⃣ Scrivi un messaggio all\'host',
     'Sub-text contestuale ("Prima di continuare, racconta...")',
     'Avatar Giuseppe + "Host dal 2026" · Textarea grande con placeholder esempio',
     'CTA NERO "Avanti" 432×40',
     '← back arrow top-left · × close top-right · step indicator dot 2/4'),
    ('3️⃣ Vuoi aggiungere un\'assicurazione di viaggio?',
     'Card "Sì, aggiungi per 17,38 €" + sub-text "Disponibile solo al momento della prenotazione"',
     'Bottone "Aggiungi" (white-bordered) · Sub-text "Ottieni rimborso 100%..." · "Cosa copre" link',
     'CTA NERO "Avanti" 432×40 (skip insurance)',
     'Click "Aggiungi" apre SUB-MODAL termini assicurazione con: long legal text + 2 footer btns "No, grazie" / "Conferma e aggiungi" (NERO)'),
    ('4️⃣ Invia una richiesta di prenotazione',
     'Card Metodo pagamento (VISA 4548 + → freccia) · Card Messaggio (preview testo + → freccia) · Card Assicurazione (last chance)',
     'Brand row VISA·MC·AMEX·PayPal·GPay · Dettagli del prezzo INLINE (4 notti + tasse + totale + Riepilogo costi link) · Disclaimer 24h + termini',
     'CTA ROSA #FF385C "Invia una richiesta di prenotazione" 432×52',
     '★ KEY: CTA cambia da NERO → ROSA (action irreversibile/conversion) · No back arrow (solo ×) · Footer Privacy·Termini·Dettagli azienda'),
    ('🔄 Backtracking',
     'Da step 2-3-4 si può tornare indietro con ← back arrow top-left',
     'Lo stato (testo messaggio, assicurazione yes/no) si conserva nel form-state',
     '—',
     'Click ✕ chiudi: torna alla pagina alloggio (esce dal checkout, perde lo stato)'),
    ('🎨 Color logic CTA',
     'Step 1-3: "Avanti" NERO #222 (action sicura/intermedia)',
     'Step 4 (final review): "Invia richiesta" ROSA #FF385C (action irreversibile · conversion)',
     '—',
     'Pattern intenzionale brand-led · il rosa è SEMPRE riservato all\'azione di conversion finale'),
    ('📍 Bottom nav durante wizard',
     'Bottom nav 5 voci (Esplora/Preferiti/Viaggi/Messaggi/Profilo) RESTA presente in tutti gli step',
     'Permette uscita laterale dal wizard senza confermare',
     '—',
     'Pattern Airbnb signature · NON tutti i wizard mantengono la navigazione globale (alcuni sono "trap-modal" full)'),
]

for r_idx, row in enumerate(flow_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx == 0 else FONT_LBL
        cell.alignment = ALIGN_L
        cell.border = BOX
        if isinstance(val, str) and val.startswith(('1️⃣', '2️⃣', '3️⃣', '4️⃣')):
            cell.fill = PatternFill('solid', start_color='dbeafe', end_color='dbeafe')
        elif isinstance(val, str) and val.startswith(('🔄', '🎨', '📍')):
            cell.fill = PatternFill('solid', start_color='fef3c7', end_color='fef3c7')


# ═════════════════════════════════════════════════════════════════════════════
# FOGLIO 8: ux-patterns-vs-livingapple (sintesi finale + adoption recommendations)
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('ux-patterns-vs-livingapple')
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 36
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 50
ws.column_dimensions['E'].width = 12

ws['A1'] = 'Airbnb checkout — Pattern UX vs LivingApple /paga (PagaClient.tsx) — adoption matrix'
ws['A1'].font = FONT_TITLE
ws.merge_cells('A1:E1')

make_table_header(ws, 3, ['#', 'Pattern Airbnb', 'Stato in LivingApple /paga', 'Adoption raccomandata', 'Priority'])

ux_rows = [
    (1, 'Layout 2-col asimmetrico (form 540 + summary 340 sticky)',
     '/paga è figlia di wizardstep2 (BookingSidebar) — ha già form sx + sidebar dx',
     'GIÀ ADOTTATO · LivingApple BookingSidebar è 380w (vs 340 Airbnb) · proporzioni simili', 'N/A'),
    (2, 'Summary card STICKY top:32 (su ultrawide)',
     'BookingSidebar è static (non sticky) attualmente',
     'INTRODURRE: position:sticky top:32px sul container BookingSidebar a >=1024px · UX premium · costo basso (1 CSS)', '🟢 ALTA'),
    (3, 'Modale Metodo di pagamento separata (5 metodi: Carta · Apple Pay · PayPal · GPay · Klarna)',
     'LivingApple ha solo Stripe (carta) + PayPal · Apple Pay/GPay non offerti',
     'VALUTARE: Stripe Elements supporta Apple Pay/Google Pay nativamente via "Payment Request Button" · costo dev medio · impatto conversion mobile alto', '🟡 MEDIA'),
    (4, 'Klarna (BNPL "Paga in 3 rate") integrato come radio in card "Scegli quando pagare"',
     'LivingApple non offre BNPL',
     'VALUTARE: Stripe ha integrazione Klarna · può aumentare prenotazioni alto-prezzo · ma serve compliance con normative italiane su credito al consumo', '🟡 MEDIA'),
    (5, 'Price transparency: "Riepilogo dei costi" modale con breakdown 4 voci + spiegazione price drop',
     'BookingSidebar mostra breakdown inline (notti×prezzo + cleaning + servizio + tasse + totale)',
     'INTRODURRE: spiegazione contestuale prezzi quando offer attive (es. "Prezzo scontato perché [reason]") · pattern brand trust', '🟢 ALTA'),
    (6, 'CTA color logic: ROSA brand SOLO per submit-finale (azioni intermedie sono NERE/grigie)',
     'LivingApple usa --color-primary BLU per tutti i CTA principali',
     'NON ADOTTARE: la decisione brand di LivingApple è coerente con palette blu (booking.com pattern) · il rosa Airbnb è brand-specific, non transferable', '🔴 NO'),
    (7, 'Modifica btn 81×32 grigio (Date · Ospiti · Metodo) — design system primary secondary',
     'LivingApple non ha pattern Modifica nel /paga (perché /paga arriva da wizardstep3, non c\'è modifica inline)',
     'VALUTARE: aggiungere link "Modifica date/ospiti" nel BookingSidebar per UX flessibile · ma rischia di rompere il funnel · da testare', '🟡 BASSA'),
    (8, 'Field grouping (5 cards form gap 24): Scegli quando pagare · Metodo · Messaggio · Assicurazione · Submit',
     'LivingApple ha sezioni: Dati ospite · Metodo pagamento · Termini · Submit',
     'COERENTE: pattern simile · Airbnb in più ha "Messaggio host" (specifico request-to-book) e "Assicurazione" (upsell) · LivingApple potrebbe testare upsell (es. cleaning premium)', '🟡 MEDIA'),
    (9, 'WIZARD MOBILE 4-step (entrata a <520) — fullscreen drawer + step indicator + CTA color change',
     'LivingApple su mobile mantiene single-page checkout (form + summary stacked)',
     'NON ADOTTARE A BREVE: il wizard mobile 4-step richiede refactor MAJOR · pattern Booking.com (LivingApple riferimento) usa single-page anche mobile · valutare metriche prima', '🔴 BASSO'),
    (10, 'Login wall: checkout richiede login Airbnb (no guest checkout)',
     'LivingApple consente guest checkout (no account)',
     'NON ADOTTARE: guest checkout è UX migliore per booking site · il login wall di Airbnb funziona perché ha 200M+ utenti già registrati · LivingApple non ha quella scala', '🔴 NO'),
    (11, 'Trust signals: NESSUN logo Stripe/secure visibile · brand stesso fa da trust',
     'LivingApple ha "Stripe Logo" (legacy) o nessun trust signal esplicito',
     'COERENTE: Airbnb non mostra Stripe logo · trust signal è il brand stesso · LivingApple costruirà brand trust nel tempo · per ora valutare se lasciare/togliere logo Stripe', '🟡 BASSA'),
    (12, 'Carte scadute mostrate ma DISABILITATE nel modale (UX trasparente)',
     'LivingApple non ha saved-card management',
     'VALUTARE FUTURO: se introdotto Stripe Customer + saved methods, mantenere carte scadute con stato disabled è UX onesta', '🟡 BASSA'),
    (13, 'Disclaimer "L\'host ha 24 ore per confermare" specifico per request-to-book',
     'LivingApple usa Beds24 conferma immediata (no 24h wait)',
     'NON APPLICABILE: LivingApple non ha host-side approval flow · il disclaimer Airbnb è per il modello marketplace (host può rifiutare)', '🔴 N/A'),
]

for r_idx, row in enumerate(ux_rows):
    for c_idx, val in enumerate(row):
        cell = ws.cell(row=5 + r_idx, column=c_idx + 1, value=val)
        cell.font = FONT_LBL_BOLD if c_idx in (0, 1) else FONT_LBL
        cell.alignment = ALIGN_C if c_idx in (0, 4) else ALIGN_L
        cell.border = BOX
        if c_idx == 4 and isinstance(val, str):
            if '🟢' in val:
                cell.fill = PatternFill('solid', start_color='d1fae5', end_color='d1fae5')
            elif '🟡' in val:
                cell.fill = PatternFill('solid', start_color='fef3c7', end_color='fef3c7')
            elif '🔴' in val:
                cell.fill = PatternFill('solid', start_color='fee2e2', end_color='fee2e2')

# Sintesi top-priority
note_row = 5 + len(ux_rows) + 2
ws.cell(row=note_row, column=1, value='TOP 3 ADOPTION HIGH PRIORITY (consigliato per /paga prossima sessione):').font = Font(name='Arial', size=11, bold=True)
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)

priorities = [
    '1. Sticky summary card (BookingSidebar position:sticky top:32 a >=1024) — 1 CSS line, alto impatto UX premium',
    '2. Price transparency contestuale (spiegazione "perché questo prezzo è scontato") — quando offer attive · ↑ trust',
    '3. Apple Pay + Google Pay via Stripe Payment Request Button — ↑ conversion mobile (~10-15% booking sites)',
    '',
    'TOP 3 NON ADOTTARE (specific to Airbnb model):',
    '1. CTA rosa #FF385C — è brand-specific Airbnb · LivingApple deve mantenere blu (Booking.com pattern coerente)',
    '2. Wizard mobile 4-step — refactor major · pattern Booking.com (riferimento) usa single-page mobile',
    '3. Login wall — Airbnb può permetterselo (200M utenti) · LivingApple deve offrire guest checkout',
]
for i, n in enumerate(priorities):
    cell = ws.cell(row=note_row + 1 + i, column=1, value=n)
    cell.font = FONT_LBL_BOLD if (i == 4) else FONT_LBL
    ws.merge_cells(start_row=note_row + 1 + i, start_column=1, end_row=note_row + 1 + i, end_column=5)


# ═════════════════════════════════════════════════════════════════════════════
# SAVE
# ═════════════════════════════════════════════════════════════════════════════
import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'docs', 'ux', 'airbnb-checkout-analysis.xlsx'))
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f'OK: {out_path} ({os.path.getsize(out_path)} bytes) · {len(wb.sheetnames)} fogli')
print(f'Fogli: {wb.sheetnames}')
